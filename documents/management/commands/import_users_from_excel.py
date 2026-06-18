import re
from pathlib import Path

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from documents.models import Department, Document, DocumentMovement, UserProfile


class Command(BaseCommand):
    help = 'Import users and profiles from an Excel workbook.'

    header_aliases = {
        'username': ['username', 'user name', 'staff id'],
        'email': ['email', 'email id', 'email id '],
        'full_name': ['full name', 'staff name', 'name'],
        'first_name': ['first_name', 'first name'],
        'last_name': ['last_name', 'last name'],
        'department': ['department', 'department full name', 'department full'],
        'department_code': ['department code'],
        'designation': ['designation', 'designation full', 'designation code'],
        'phone': ['phone', 'mobile number', 'mobile', 'phone number'],
        'role': ['role'],
        'group': ['group', 'groups'],
        'is_active': ['is_active', 'active'],
    }
    role_group_map = {
        'admin': 'Admin',
        'receiving_desk': 'Receiving Desk',
        'hr': 'HR',
        'adt': 'ADT',
        'ops': 'OPS',
        'management': 'Management',
    }
    role_aliases = {
        'admin': 'admin',
        'administrator': 'admin',
        'receiving desk': 'receiving_desk',
        'receiving_desk': 'receiving_desk',
        'hr': 'hr',
        'hr&a': 'hr',
        'human resource': 'hr',
        'adt': 'adt',
        'analytics digital transformation': 'adt',
        'analytics & digital transformation': 'adt',
        'ops': 'ops',
        'operations': 'ops',
        'management': 'management',
    }
    allowed_group_names = set(role_group_map.values())
    staff_id_username_pattern = re.compile(r'^\d{10}$')

    def add_arguments(self, parser):
        parser.add_argument('excel_path', help='Path to user Excel file')
        parser.add_argument(
            '--sheet',
            default='Sheet1',
            help='Worksheet name to import. Defaults to Sheet1.',
        )
        parser.add_argument(
            '--default-password',
            default='',
            help='Optional temporary password for created users.',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview import without writing users.',
        )
        parser.add_argument(
            '--replace-name-usernames',
            action='store_true',
            help=(
                'Delete old imported users with name-based usernames, reimport '
                'staff-ID users, and restore document links by email.'
            ),
        )

    def handle(self, *args, **options):
        excel_path = Path(options['excel_path'])
        if not excel_path.exists():
            raise CommandError(f'Excel file not found: {excel_path}')

        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        sheet_name = options['sheet']
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f'Sheet "{sheet_name}" not found. Available sheets: '
                + ', '.join(workbook.sheetnames)
            )

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            raise CommandError('Selected sheet is empty.')

        column_map = self.get_column_map(header_row)
        if 'email' not in column_map and 'username' not in column_map:
            raise CommandError('Excel must contain at least email or username column.')

        import_rows = []
        skipped_blank_rows = 0

        for row_number, row in enumerate(rows, start=2):
            data = self.clean_row(row, column_map)
            if not any(data.values()):
                skipped_blank_rows += 1
                continue
            import_rows.append((row_number, data))

        workbook.close()

        workbook_emails = self.workbook_emails(import_rows)
        replacement_users = []
        reference_snapshots = []
        ignored_user_ids = set()

        if options['replace_name_usernames']:
            replacement_users = self.replacement_candidates(workbook_emails)
            ignored_user_ids = {user.pk for user in replacement_users}
            reference_snapshots = self.capture_user_references(replacement_users)
            self.validate_replacement_references(reference_snapshots, workbook_emails)

        if options['replace_name_usernames'] and not options['dry_run']:
            with transaction.atomic():
                deleted_count = self.delete_replacement_users(replacement_users)
                summary = self.import_rows(
                    import_rows,
                    skipped_blank_rows,
                    default_password=options['default_password'],
                    dry_run=False,
                )
                if summary['errors']:
                    raise CommandError(
                        'Replacement import stopped because one or more rows failed.'
                    )
                restored_count = self.restore_user_references(reference_snapshots)
        else:
            deleted_count = len(replacement_users) if options['replace_name_usernames'] else 0
            restored_count = len(reference_snapshots) if options['replace_name_usernames'] else 0
            summary = self.import_rows(
                import_rows,
                skipped_blank_rows,
                default_password=options['default_password'],
                dry_run=options['dry_run'],
                ignored_user_ids=ignored_user_ids,
            )

        label = 'User import preview complete.' if options['dry_run'] else 'User import complete.'
        self.stdout.write(self.style.SUCCESS(label))
        self.stdout.write(f"Created: {summary['created']}")
        self.stdout.write(f"Updated: {summary['updated']}")
        self.stdout.write(f"Skipped: {summary['skipped']}")
        self.stdout.write(f"Errors: {summary['errors']}")
        if options['replace_name_usernames']:
            replacement_label = 'Would replace' if options['dry_run'] else 'Replaced'
            restore_label = 'Would restore' if options['dry_run'] else 'Restored'
            self.stdout.write(f'{replacement_label} name-based users: {deleted_count}')
            self.stdout.write(f'{restore_label} document links: {restored_count}')

        if not options['default_password']:
            self.stdout.write(
                self.style.WARNING(
                    'Created users have unusable passwords unless they already had a password. '
                    'Use --default-password for first-time login access.'
                )
            )

    def import_rows(
        self,
        import_rows,
        skipped_blank_rows,
        default_password='',
        dry_run=False,
        ignored_user_ids=None,
    ):
        summary = {
            'created': 0,
            'updated': 0,
            'skipped': skipped_blank_rows,
            'errors': 0,
        }

        for row_number, data in import_rows:
            try:
                result = self.import_row(
                    data,
                    default_password=default_password,
                    dry_run=dry_run,
                    ignored_user_ids=ignored_user_ids,
                )
            except ValueError as exc:
                summary['errors'] += 1
                self.stderr.write(f'Row {row_number} skipped: {exc}')
                continue

            summary[result] += 1

        return summary

    def get_column_map(self, header_row):
        normalized_headers = {
            self.normalize_header(header): index
            for index, header in enumerate(header_row)
            if header is not None
        }
        column_map = {}

        for canonical, aliases in self.header_aliases.items():
            for alias in aliases:
                index = normalized_headers.get(self.normalize_header(alias))
                if index is not None:
                    column_map[canonical] = index
                    break

        return column_map

    def clean_row(self, row, column_map):
        cleaned = {}
        for key, index in column_map.items():
            value = row[index] if index < len(row) else ''
            cleaned[key] = self.clean_cell(value)
        return cleaned

    def import_row(
        self,
        data,
        default_password='',
        dry_run=False,
        ignored_user_ids=None,
    ):
        email = data.get('email', '').lower()
        username = data.get('username') or self.username_from_email(email)
        username = self.normalize_username(username)

        if not username:
            raise ValueError('username or email is required.')

        user = self.find_existing_user(username, email, ignored_user_ids)

        first_name, last_name = self.name_parts(data)
        department = self.get_department(
            data.get('department', ''),
            data.get('department_code', ''),
            dry_run,
        )
        role = self.normalize_role(data.get('role', ''))
        group_names = self.group_names_for(data.get('group', ''), role)
        is_active = self.parse_bool(data.get('is_active', ''), default=True)

        if dry_run:
            return 'updated' if user else 'created'

        if user:
            result = 'updated'
        else:
            user_model = get_user_model()
            user = user_model(username=username)
            if default_password:
                user.set_password(default_password)
            else:
                user.set_unusable_password()
            result = 'created'

        user.email = email
        user.first_name = first_name
        user.last_name = last_name
        user.is_active = is_active
        user.save()

        for group_name in group_names:
            group, _ = Group.objects.get_or_create(name=group_name)
            user.groups.add(group)

        UserProfile.objects.update_or_create(
            user=user,
            defaults={
                'department': department,
                'role': role,
                'designation': data.get('designation', ''),
                'phone': data.get('phone', ''),
            },
        )

        return result

    def find_existing_user(self, username, email, ignored_user_ids=None):
        user_model = get_user_model()
        users = user_model.objects.all()

        if ignored_user_ids:
            users = users.exclude(pk__in=ignored_user_ids)

        user = users.filter(username__iexact=username).first()
        if not user and email:
            user = users.filter(email__iexact=email).first()

        return user

    def workbook_emails(self, import_rows):
        return {
            data.get('email', '').lower()
            for _, data in import_rows
            if data.get('email', '')
        }

    def replacement_candidates(self, workbook_emails):
        user_model = get_user_model()
        candidates = user_model.objects.filter(
            profile__isnull=False,
            is_superuser=False,
            groups__isnull=True,
        ).distinct()

        return [
            user
            for user in candidates
            if (
                user.email and
                user.email.lower() in workbook_emails and
                not self.is_staff_id_username(user.username)
            )
        ]

    def capture_user_references(self, users):
        user_ids = [user.pk for user in users]
        if not user_ids:
            return []

        reference_fields = [
            (Document, 'designated_person'),
            (Document, 'first_boss'),
            (Document, 'created_by'),
            (DocumentMovement, 'performed_by'),
        ]
        references = []

        for model, field_name in reference_fields:
            filter_kwargs = {f'{field_name}_id__in': user_ids}
            for obj in model.objects.filter(**filter_kwargs).select_related(field_name):
                user = getattr(obj, field_name)
                references.append({
                    'model': model,
                    'pk': obj.pk,
                    'field_name': field_name,
                    'email': (user.email or '').lower(),
                })

        return references

    def validate_replacement_references(self, references, workbook_emails):
        missing_emails = sorted({
            reference['email']
            for reference in references
            if not reference['email'] or reference['email'] not in workbook_emails
        })

        if missing_emails:
            raise CommandError(
                'Replacement import stopped because referenced user email(s) '
                'are missing from the workbook: '
                + ', '.join(missing_emails)
            )

    def delete_replacement_users(self, users):
        user_ids = [user.pk for user in users]
        if not user_ids:
            return 0

        get_user_model().objects.filter(pk__in=user_ids).delete()
        return len(user_ids)

    def restore_user_references(self, references):
        if not references:
            return 0

        emails = {reference['email'] for reference in references}
        users_by_email = {
            user.email.lower(): user
            for user in get_user_model().objects.filter(email__in=emails)
            if user.email
        }

        for reference in references:
            user = users_by_email.get(reference['email'])
            if not user:
                raise CommandError(
                    'Replacement import stopped because a referenced user '
                    f'was not recreated: {reference["email"]}'
                )

            reference['model'].objects.filter(pk=reference['pk']).update(
                **{reference['field_name']: user}
            )

        return len(references)

    def get_department(self, name, code, dry_run=False):
        if not name:
            return None

        department = Department.objects.filter(name__iexact=name).first()
        if department or dry_run:
            return department

        department_code = code or self.department_code_from_name(name)
        department_code = self.unique_department_code(department_code)
        return Department.objects.create(
            name=name,
            code=department_code,
            is_active=True,
        )

    def name_parts(self, data):
        first_name = data.get('first_name', '')
        last_name = data.get('last_name', '')
        full_name = data.get('full_name', '')

        if first_name or last_name or not full_name:
            return first_name, last_name

        parts = full_name.split()
        if len(parts) == 1:
            return parts[0], ''

        return ' '.join(parts[:-1]), parts[-1]

    def group_names_for(self, group_value, role):
        group_names = set()

        if role:
            group_names.add(self.role_group_map[role])

        for raw_group in re.split(r'[,;/|]', group_value or ''):
            group_name = raw_group.strip()
            if group_name in self.allowed_group_names:
                group_names.add(group_name)

        return sorted(group_names)

    def normalize_role(self, value):
        normalized = self.normalize_header(value).replace('-', ' ')
        if not normalized:
            return ''
        role = self.role_aliases.get(normalized)
        if role:
            return role
        if normalized in self.role_group_map:
            return normalized
        return ''

    def normalize_username(self, value):
        value = (value or '').strip()
        if '@' in value:
            value = value.split('@', 1)[0]
        value = value.lower()
        value = re.sub(r'\s+', '.', value)
        value = re.sub(r'[^a-z0-9.@_+-]', '', value)
        return value[:150]

    def is_staff_id_username(self, username):
        return bool(self.staff_id_username_pattern.fullmatch(username or ''))

    def username_from_email(self, email):
        if not email or '@' not in email:
            return ''
        return email.split('@', 1)[0]

    def department_code_from_name(self, name):
        words = re.findall(r'[A-Za-z0-9]+', name)
        if words:
            code = ''.join(word[0] for word in words[:8]).upper()
        else:
            code = name[:8].upper()
        return (code or 'DEPT')[:20]

    def unique_department_code(self, code):
        base_code = (code or 'DEPT')[:20]
        candidate = base_code
        counter = 2

        while Department.objects.filter(code__iexact=candidate).exists():
            suffix = str(counter)
            candidate = f'{base_code[:20 - len(suffix)]}{suffix}'
            counter += 1

        return candidate

    def parse_bool(self, value, default=True):
        if value == '':
            return default
        if isinstance(value, bool):
            return value
        normalized = self.normalize_header(value)
        if normalized in {'1', 'true', 'yes', 'y', 'active'}:
            return True
        if normalized in {'0', 'false', 'no', 'n', 'inactive'}:
            return False
        return default

    def normalize_header(self, value):
        return ' '.join(str(value or '').strip().casefold().split())

    def clean_cell(self, value):
        if value is None:
            return ''
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()
