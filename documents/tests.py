import json
import os
import tempfile
from datetime import timedelta
from decimal import Decimal
from io import BytesIO
from unittest import mock

from django import forms
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core import mail
from django.core.management import call_command
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from openpyxl import Workbook, load_workbook

from .forms import DocumentForm, DocumentMovementForm, UserDelegationRuleForm
from .models import (
    Area,
    Branch,
    CourierRate,
    Department,
    Division,
    Document,
    DocumentDelegation,
    DocumentMovement,
    Region,
    UserDelegationRule,
    UserProfile,
)
from .views import can_edit_documents, movement_action_from_status


@override_settings(
    EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
    DEFAULT_FROM_EMAIL='receiving@example.org',
)
class HeadOfficeNotificationWorkflowTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.department = Department.objects.create(name='HR', code='HR')
        self.ops_department = Department.objects.create(name='OPS', code='OPS')
        self.division = Division.objects.create(name='Central East Division')
        self.region = Region.objects.create(
            division=self.division,
            name='Dhaka South Region',
        )
        self.area = Area.objects.create(
            region=self.region,
            name='Rampura Area',
        )
        self.branch = Branch.objects.create(
            area=self.area,
            name='Banasree Branch',
        )

        self.admin_group, _ = Group.objects.get_or_create(name='Admin')
        self.receiving_group, _ = Group.objects.get_or_create(name='Receiving Desk')
        self.management_group, _ = Group.objects.get_or_create(name='Management')

        self.receiving_user = user_model.objects.create_user(
            username='polu',
            password='password',
            email='polu@example.org',
        )
        self.receiving_user.groups.add(self.receiving_group)

        self.management_user = user_model.objects.create_user(
            username='manager',
            password='password',
            email='manager@example.org',
        )
        self.management_user.groups.add(self.management_group)

        self.designated_person = user_model.objects.create_user(
            username='recipient',
            password='password',
            email='recipient@example.org',
        )
        self.delegated_receiver = user_model.objects.create_user(
            username='delegate',
            password='password',
            email='delegate@example.org',
        )
        self.unrelated_user = user_model.objects.create_user(
            username='unrelated',
            password='password',
            email='unrelated@example.org',
        )
        self.first_boss = user_model.objects.create_user(
            username='boss',
            password='password',
            email='boss@example.org',
        )

    def create_document(self, **overrides):
        defaults = {
            'entry_type': 'inward',
            'subject': 'Policy memo',
            'current_department': self.department,
            'destination_department': self.department,
            'created_by': self.receiving_user,
        }
        defaults.update(overrides)
        return Document.objects.create(**defaults)

    def create_delegation(self, document, **overrides):
        defaults = {
            'document': document,
            'original_recipient': document.designated_person,
            'delegated_recipient': self.delegated_receiver,
            'reason': 'Original recipient is on leave.',
            'delegated_by': document.designated_person,
        }
        defaults.update(overrides)
        return DocumentDelegation.objects.create(**defaults)

    def post_forward_update(self, document):
        return self.client.post(
            reverse('document_update_movement', args=[document.pk]),
            {
                'new_status': 'forwarded',
                'to_department': self.ops_department.pk,
                'remarks': 'Physical copy forwarded to OPS.',
            },
        )

    def test_receiving_desk_can_edit_and_management_cannot(self):
        self.assertTrue(can_edit_documents(self.receiving_user))
        self.assertFalse(can_edit_documents(self.management_user))

        self.client.login(username='polu', password='password')
        self.assertEqual(self.client.get(reverse('document_create')).status_code, 200)

        self.client.logout()
        self.client.login(username='manager', password='password')
        self.assertEqual(self.client.get(reverse('document_create')).status_code, 403)

    def test_document_form_accepts_department_only_recipient_information(self):
        form = DocumentForm(data={
            'entry_type': 'inward',
            'document_type': 'Memo',
            'subject': 'Department-only memo',
            'reference_no': 'REF-001',
            'addressed_to_name': '',
            'addressed_to_designation': 'Head of OPS',
            'notification_required': 'on',
            'source_division': self.division.pk,
            'source_region': self.region.pk,
            'source_area': self.area.pk,
            'source_branch': self.branch.pk,
            'source_department': '',
            'source_zone': '',
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'received_date': '2026-06-17',
            'sent_date': '',
            'priority': 'normal',
            'status': 'physical_received',
            'remarks': '',
        })

        self.assertTrue(form.is_valid(), form.errors)

    def test_document_form_allows_blank_subject_with_required_inward_date(self):
        form = DocumentForm(data={
            'entry_type': 'inward',
            'document_type': 'Memo',
            'subject': '',
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'received_date': '2026-06-17',
            'priority': 'normal',
            'status': 'physical_received',
        })

        self.assertTrue(form.is_valid(), form.errors)
        document = form.save()
        self.assertEqual(document.subject, '')

    def test_document_form_requires_received_date_for_inward_documents(self):
        form = DocumentForm(data={
            'entry_type': 'inward',
            'document_type': 'Memo',
            'subject': '',
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'received_date': '',
            'sent_date': '2026-06-17',
            'priority': 'normal',
            'status': 'physical_received',
        })

        self.assertFalse(form.is_valid())
        self.assertEqual(
            form.errors['received_date'],
            ['Received Date is required for inward documents.'],
        )

    def test_document_form_requires_sent_date_for_outward_documents(self):
        form = DocumentForm(data={
            'entry_type': 'outward',
            'document_type': 'Memo',
            'subject': '',
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'received_date': '2026-06-17',
            'sent_date': '',
            'priority': 'normal',
            'status': 'physical_received',
        })

        self.assertFalse(form.is_valid())
        self.assertEqual(
            form.errors['sent_date'],
            ['Sent Date is required for outward documents.'],
        )

    def test_document_form_populates_selected_courier_rate_values(self):
        courier_rate = CourierRate.objects.create(
            description='ছোট খামের চিঠি',
            quantity='0-500 গ্রাম',
            amount=Decimal('25.00'),
        )

        form = DocumentForm(data={
            'entry_type': 'outward',
            'document_type': 'Memo',
            'subject': 'Courier memo',
            'notification_required': 'on',
            'source_division': self.division.pk,
            'source_region': self.region.pk,
            'source_area': self.area.pk,
            'source_branch': self.branch.pk,
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'courier_rate': courier_rate.pk,
            'sent_date': '2026-06-17',
            'outward_description': '',
            'outward_quantity': '',
            'outward_amount': '',
            'outward_is_manual': '',
            'priority': 'normal',
            'status': 'physical_received',
        })

        self.assertTrue(form.is_valid(), form.errors)
        document = form.save()

        self.assertEqual(document.courier_rate, courier_rate)
        self.assertEqual(document.outward_description, courier_rate.description)
        self.assertEqual(document.outward_quantity, courier_rate.quantity)
        self.assertEqual(document.outward_amount, courier_rate.amount)
        self.assertFalse(document.outward_is_manual)

    def test_document_form_manual_courier_values_clear_selected_rate(self):
        courier_rate = CourierRate.objects.create(
            description='বড় খামের চিঠি',
            quantity='0-500 গ্রাম',
            amount=Decimal('35.00'),
        )

        form = DocumentForm(data={
            'entry_type': 'outward',
            'document_type': 'Memo',
            'subject': 'Manual courier memo',
            'notification_required': 'on',
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'courier_rate': courier_rate.pk,
            'sent_date': '2026-06-17',
            'outward_description': 'Other delivery',
            'outward_quantity': '1 parcel',
            'outward_amount': '10.50',
            'outward_is_manual': 'on',
            'priority': 'normal',
            'status': 'physical_received',
        })

        self.assertTrue(form.is_valid(), form.errors)
        document = form.save()

        self.assertIsNone(document.courier_rate)
        self.assertEqual(document.outward_description, 'Other delivery')
        self.assertEqual(document.outward_quantity, '1 parcel')
        self.assertEqual(document.outward_amount, Decimal('10.50'))
        self.assertTrue(document.outward_is_manual)

    def test_document_form_rejects_negative_courier_amount(self):
        form = DocumentForm(data={
            'entry_type': 'outward',
            'document_type': 'Memo',
            'subject': 'Negative cost memo',
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'sent_date': '2026-06-17',
            'outward_amount': '-1.00',
            'priority': 'normal',
            'status': 'physical_received',
        })

        self.assertFalse(form.is_valid())
        self.assertIn('outward_amount', form.errors)

    def test_source_hierarchy_fields_are_available_on_document_form(self):
        form = DocumentForm()

        self.assertIn('source_type', form.fields)
        self.assertIn('external_organization_type', form.fields)
        self.assertIn('external_organization_name', form.fields)
        self.assertIn('external_branch_name', form.fields)
        self.assertIn('source_division', form.fields)
        self.assertIn('source_region', form.fields)
        self.assertIn('source_area', form.fields)
        self.assertIn('source_branch', form.fields)
        self.assertNotIn('source_zone', form.fields)

    def test_external_organization_type_choices_include_government_offices(self):
        form = DocumentForm()
        choices = dict(form.fields['external_organization_type'].choices)

        self.assertEqual(choices['government_offices'], 'Government Offices')

    def test_document_form_accepts_external_pksf_without_organization_details(self):
        form = DocumentForm(data={
            'entry_type': 'inward',
            'document_type': 'Memo',
            'subject': 'External PKSF memo',
            'source_type': 'external',
            'external_organization_type': 'pksf',
            'external_organization_name': '',
            'external_branch_name': '',
            'source_division': self.division.pk,
            'source_region': self.region.pk,
            'source_area': self.area.pk,
            'source_branch': self.branch.pk,
            'source_department': self.department.pk,
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'received_date': '2026-06-17',
            'priority': 'normal',
            'status': 'physical_received',
        })

        self.assertTrue(form.is_valid(), form.errors)
        document = form.save()

        self.assertEqual(document.source_type, 'external')
        self.assertEqual(document.external_organization_type, 'pksf')
        self.assertIsNone(document.source_division)
        self.assertIsNone(document.source_department)

    def test_document_form_requires_bank_organization_name_and_branch(self):
        form = DocumentForm(data={
            'entry_type': 'inward',
            'document_type': 'Memo',
            'subject': 'External bank memo',
            'source_type': 'external',
            'external_organization_type': 'bank',
            'external_organization_name': '',
            'external_branch_name': '',
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'received_date': '2026-06-17',
            'priority': 'normal',
            'status': 'physical_received',
        })

        self.assertFalse(form.is_valid())
        self.assertIn('external_organization_name', form.errors)
        self.assertIn('external_branch_name', form.errors)

        valid_form = DocumentForm(data={
            'entry_type': 'inward',
            'document_type': 'Memo',
            'subject': 'External bank memo',
            'source_type': 'external',
            'external_organization_type': 'bank',
            'external_organization_name': 'City Bank',
            'external_branch_name': 'Gulshan',
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'received_date': '2026-06-17',
            'priority': 'normal',
            'status': 'physical_received',
        })

        self.assertTrue(valid_form.is_valid(), valid_form.errors)

    def test_document_form_accepts_government_offices_for_inward_and_outward(self):
        inward_form = DocumentForm(data={
            'entry_type': 'inward',
            'document_type': 'Memo',
            'subject': 'External government memo',
            'source_type': 'external',
            'external_organization_type': 'government_offices',
            'external_organization_name': 'Ministry of Finance',
            'external_branch_name': 'Dhaka',
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'received_date': '2026-06-17',
            'priority': 'normal',
            'status': 'physical_received',
        })
        outward_form = DocumentForm(data={
            'entry_type': 'outward',
            'document_type': 'Memo',
            'subject': 'Outbound government memo',
            'source_type': 'external',
            'external_organization_type': 'government_offices',
            'external_organization_name': 'Ministry of Finance',
            'external_branch_name': 'Dhaka',
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'sent_date': '2026-06-17',
            'priority': 'normal',
            'status': 'physical_received',
        })

        self.assertTrue(inward_form.is_valid(), inward_form.errors)
        self.assertTrue(outward_form.is_valid(), outward_form.errors)

    def test_document_form_requires_government_office_name_and_branch(self):
        form = DocumentForm(data={
            'entry_type': 'inward',
            'document_type': 'Memo',
            'subject': 'External government memo',
            'source_type': 'external',
            'external_organization_type': 'government_offices',
            'external_organization_name': '',
            'external_branch_name': '',
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'received_date': '2026-06-17',
            'priority': 'normal',
            'status': 'physical_received',
        })

        self.assertFalse(form.is_valid())
        self.assertIn('external_organization_name', form.errors)
        self.assertIn('external_branch_name', form.errors)

    def test_temporary_delegation_rule_validates_date_range_and_backup_receiver(self):
        today = timezone.now().date()
        invalid_date_form = UserDelegationRuleForm(
            user=self.designated_person,
            data={
                'backup_receiver': self.delegated_receiver.pk,
                'start_date': today,
                'end_date': today - timedelta(days=1),
                'reason': 'Leave.',
            },
        )
        same_user_form = UserDelegationRuleForm(
            user=self.designated_person,
            data={
                'backup_receiver': self.designated_person.pk,
                'start_date': today,
                'end_date': today,
                'reason': 'Leave.',
            },
        )

        self.assertFalse(invalid_date_form.is_valid())
        self.assertIn('end_date', invalid_date_form.errors)
        self.assertFalse(same_user_form.is_valid())
        self.assertIn('backup_receiver', same_user_form.errors)

    def test_source_hierarchy_querysets_follow_selected_parent(self):
        form = DocumentForm(data={
            'entry_type': 'inward',
            'document_type': 'Memo',
            'subject': 'Hierarchy memo',
            'source_division': self.division.pk,
            'source_region': self.region.pk,
            'source_area': self.area.pk,
            'source_branch': self.branch.pk,
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'received_date': '2026-06-17',
            'priority': 'normal',
            'status': 'physical_received',
        })

        self.assertIn(self.region, form.fields['source_region'].queryset)
        self.assertIn(self.area, form.fields['source_area'].queryset)
        self.assertIn(self.branch, form.fields['source_branch'].queryset)

    def test_document_type_uses_dropdown_choices(self):
        form = DocumentForm()

        self.assertIsInstance(form.fields['document_type'].widget, forms.Select)
        self.assertEqual(
            form.fields['document_type'].widget.attrs['data-searchable-select'],
            'true',
        )
        self.assertIn(('Memo', 'Memo'), form.fields['document_type'].choices)
        self.assertIn(('Complaint', 'Complaint'), form.fields['document_type'].choices)

    def test_person_dropdowns_are_searchable(self):
        form = DocumentForm()

        self.assertEqual(
            form.fields['designated_person'].widget.attrs['data-searchable-select'],
            'true',
        )
        self.assertEqual(
            form.fields['first_boss'].widget.attrs['data-searchable-select'],
            'true',
        )

    def test_user_info_api_returns_basic_staff_information(self):
        UserProfile.objects.create(
            user=self.designated_person,
            department=self.ops_department,
            designation='Senior Officer',
        )

        self.client.login(username='polu', password='password')
        response = self.client.get(
            reverse('get_user_info'),
            {'user_id': self.designated_person.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            json.loads(response.content),
            {
                'id': self.designated_person.pk,
                'name': self.designated_person.username,
                'designation': 'Senior Officer',
                'department': self.ops_department.name,
            },
        )

    def test_user_info_api_requires_login(self):
        response = self.client.get(
            reverse('get_user_info'),
            {'user_id': self.designated_person.pk},
        )

        self.assertEqual(response.status_code, 302)

    def test_document_form_keeps_selected_current_department(self):
        UserProfile.objects.create(
            user=self.designated_person,
            department=self.ops_department,
            designation='Senior Officer',
        )

        form = DocumentForm(data={
            'entry_type': 'inward',
            'document_type': 'Memo',
            'subject': 'Assigned memo',
            'designated_person': self.designated_person.pk,
            'destination_department': self.ops_department.pk,
            'current_department': self.department.pk,
            'received_date': '2026-06-17',
            'priority': 'normal',
            'status': 'physical_received',
        })

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['current_department'], self.department)

    def test_new_document_defaults_current_department_to_receiving_desk(self):
        receiving_department = Department.objects.create(
            name='Receiving Desk',
            code='RECEIVING',
        )

        self.client.login(username='polu', password='password')
        response = self.client.get(reverse('document_create'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context['form'].initial['current_department'],
            receiving_department.pk,
        )
        self.assertEqual(
            response.context['receiving_desk_department_id'],
            receiving_department.pk,
        )

    def test_new_document_exposes_user_department_default_for_outward(self):
        UserProfile.objects.create(
            user=self.receiving_user,
            department=self.ops_department,
        )

        self.client.login(username='polu', password='password')
        response = self.client.get(reverse('document_create'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['user_department_id'], self.ops_department.pk)

    def test_document_form_status_dropdown_uses_limited_choices(self):
        form = DocumentForm()

        self.assertEqual(
            list(form.fields['status'].choices),
            [
                ('created', 'Entry Created'),
                ('physical_received', 'Physical Document Received'),
                ('forwarded', 'Forwarded'),
                ('in_progress', 'In Progress'),
                ('closed', 'Closed'),
                ('missing_physical_copy', 'Missing Physical Copy'),
            ],
        )

    def test_current_department_is_labeled_receiving_department(self):
        document_form = DocumentForm()
        movement_form = DocumentMovementForm()

        self.assertEqual(document_form.fields['current_department'].label, 'Current / Receiving Department')
        self.assertEqual(movement_form.fields['to_department'].label, 'Forward / Receiving Department')
        self.assertNotIn('action', movement_form.fields)

    def test_forwarding_sends_to_designated_person_and_ccs_first_boss(self):
        document = self.create_document(
            reference_no='REF-CC',
            designated_person=self.designated_person,
            first_boss=self.first_boss,
        )

        self.client.login(username='polu', password='password')
        response = self.post_forward_update(document)

        self.assertRedirects(response, reverse('document_detail', args=[document.pk]))
        document.refresh_from_db()
        self.assertEqual(document.status, 'forwarded')
        self.assertEqual(document.current_department, self.ops_department)
        self.assertEqual(document.destination_department, self.ops_department)
        self.assertTrue(document.notification_sent)
        self.assertEqual(document.notification_error, '')
        self.assertEqual(DocumentMovement.objects.get(document=document).action, 'forwarded')
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ['recipient@example.org'])
        self.assertEqual(mail.outbox[0].cc, ['boss@example.org'])

    def test_status_update_derives_non_forwarded_movement_action(self):
        document = self.create_document()

        self.client.login(username='polu', password='password')
        response = self.client.post(
            reverse('document_update_movement', args=[document.pk]),
            {
                'new_status': 'in_progress',
                'to_department': '',
                'remarks': 'Work started.',
            },
        )

        self.assertRedirects(response, reverse('document_detail', args=[document.pk]))
        document.refresh_from_db()
        movement = DocumentMovement.objects.get(document=document)
        self.assertEqual(document.status, 'in_progress')
        self.assertEqual(movement.action, 'status_updated')

    def test_movement_action_is_derived_from_status(self):
        self.assertEqual(movement_action_from_status('forwarded'), 'forwarded')
        self.assertEqual(movement_action_from_status('physical_received'), 'physical_received')
        self.assertEqual(movement_action_from_status('received_by_department'), 'received')
        self.assertEqual(movement_action_from_status('closed'), 'closed')
        self.assertEqual(movement_action_from_status('returned'), 'returned')
        self.assertEqual(movement_action_from_status('in_progress'), 'status_updated')

    def test_forwarding_sends_only_to_designated_person_without_first_boss(self):
        document = self.create_document(designated_person=self.designated_person)

        self.client.login(username='polu', password='password')
        response = self.post_forward_update(document)

        self.assertRedirects(response, reverse('document_detail', args=[document.pk]))
        document.refresh_from_db()
        self.assertTrue(document.notification_sent)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ['recipient@example.org'])
        self.assertEqual(mail.outbox[0].cc, [])

    def test_forwarding_without_designated_person_is_allowed_when_notification_not_required(self):
        document = self.create_document(notification_required=False)

        self.client.login(username='polu', password='password')
        response = self.post_forward_update(document)

        self.assertRedirects(response, reverse('document_detail', args=[document.pk]))
        document.refresh_from_db()
        self.assertEqual(document.status, 'forwarded')
        self.assertFalse(document.notification_sent)
        self.assertEqual(document.notification_error, '')
        self.assertEqual(len(mail.outbox), 0)

    def test_forwarding_requires_designated_person_email_when_notification_required(self):
        document = self.create_document()

        self.client.login(username='polu', password='password')
        response = self.post_forward_update(document)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Forwarding requires a designated person with an email')
        document.refresh_from_db()
        self.assertEqual(document.status, 'created')
        self.assertEqual(DocumentMovement.objects.filter(document=document).count(), 0)
        self.assertEqual(len(mail.outbox), 0)

    def test_designated_person_can_close_assigned_document(self):
        document = self.create_document(
            designated_person=self.designated_person,
            status='in_progress',
        )

        self.client.login(username='recipient', password='password')
        detail_response = self.client.get(reverse('document_detail', args=[document.pk]))
        response = self.client.post(
            reverse('document_update_movement', args=[document.pk]),
            {
                'new_status': 'closed',
                'to_department': self.ops_department.pk,
                'remarks': 'Action completed.',
            },
        )

        self.assertContains(detail_response, 'Close Assigned File')
        self.assertRedirects(response, reverse('document_detail', args=[document.pk]))
        document.refresh_from_db()
        movement = DocumentMovement.objects.get(document=document)
        self.assertEqual(document.status, 'closed')
        self.assertEqual(document.current_department, self.department)
        self.assertEqual(movement.action, 'closed')
        self.assertEqual(movement.performed_by, self.designated_person)

    def test_designated_person_cannot_forward_assigned_document(self):
        document = self.create_document(
            designated_person=self.designated_person,
            status='in_progress',
        )

        self.client.login(username='recipient', password='password')
        response = self.client.post(
            reverse('document_update_movement', args=[document.pk]),
            {
                'new_status': 'forwarded',
                'to_department': self.ops_department.pk,
                'remarks': 'Try to forward.',
            },
        )

        self.assertEqual(response.status_code, 200)
        document.refresh_from_db()
        self.assertEqual(document.status, 'in_progress')
        self.assertEqual(DocumentMovement.objects.filter(document=document).count(), 0)

    def test_original_designated_person_can_access_delegate_form(self):
        document = self.create_document(
            designated_person=self.designated_person,
            status='in_progress',
        )

        self.client.login(username='recipient', password='password')
        response = self.client.get(reverse('document_delegate_receipt', args=[document.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Delegate Receipt')

    def test_unrelated_user_cannot_delegate_document(self):
        document = self.create_document(
            designated_person=self.designated_person,
            status='in_progress',
        )

        self.client.login(username='unrelated', password='password')
        response = self.client.get(reverse('document_delegate_receipt', args=[document.pk]))

        self.assertEqual(response.status_code, 403)

    def test_delegate_receipt_requires_reason(self):
        document = self.create_document(
            designated_person=self.designated_person,
            status='in_progress',
        )

        self.client.login(username='recipient', password='password')
        response = self.client.post(
            reverse('document_delegate_receipt', args=[document.pk]),
            {
                'delegated_recipient': self.delegated_receiver.pk,
                'reason': '',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Reason for delegation is required.')
        self.assertFalse(DocumentDelegation.objects.filter(document=document).exists())

    def test_delegate_receipt_creates_active_delegation_and_movement(self):
        document = self.create_document(
            designated_person=self.designated_person,
            status='in_progress',
        )

        self.client.login(username='recipient', password='password')
        response = self.client.post(
            reverse('document_delegate_receipt', args=[document.pk]),
            {
                'delegated_recipient': self.delegated_receiver.pk,
                'reason': 'Recipient is on leave.',
            },
        )

        self.assertRedirects(response, reverse('document_detail', args=[document.pk]))
        delegation = DocumentDelegation.objects.get(document=document)
        movement = DocumentMovement.objects.get(document=document, action='delegated')
        self.assertTrue(delegation.is_active)
        self.assertEqual(delegation.original_recipient, self.designated_person)
        self.assertEqual(delegation.delegated_recipient, self.delegated_receiver)
        self.assertEqual(delegation.delegated_by, self.designated_person)
        self.assertIn('Receipt delegated from', movement.remarks)
        self.assertIn('Recipient is on leave.', movement.remarks)

    def test_closed_document_cannot_be_delegated(self):
        document = self.create_document(
            designated_person=self.designated_person,
            status='closed',
        )

        self.client.login(username='recipient', password='password')
        response = self.client.get(reverse('document_delegate_receipt', args=[document.pk]))

        self.assertEqual(response.status_code, 403)

    def test_delegate_form_suggests_active_temporary_rule_backup_receiver(self):
        document = self.create_document(
            designated_person=self.designated_person,
            status='in_progress',
        )
        today = timezone.now().date()
        UserDelegationRule.objects.create(
            user=self.designated_person,
            backup_receiver=self.delegated_receiver,
            start_date=today,
            end_date=today + timedelta(days=2),
            reason='Annual leave.',
            created_by=self.designated_person,
        )

        self.client.login(username='recipient', password='password')
        response = self.client.get(reverse('document_delegate_receipt', args=[document.pk]))

        self.assertEqual(
            response.context['form'].initial['delegated_recipient'],
            self.delegated_receiver.pk,
        )

    def test_delegated_recipient_can_see_delegated_document_in_my_assigned(self):
        document = self.create_document(
            designated_person=self.designated_person,
            status='in_progress',
        )
        self.create_delegation(document)

        self.client.login(username='delegate', password='password')
        response = self.client.get(reverse('my_assigned_documents'))

        self.assertContains(response, document.tracking_id)

    def test_delegated_recipient_can_view_and_close_delegated_document(self):
        document = self.create_document(
            designated_person=self.designated_person,
            status='in_progress',
        )
        self.create_delegation(document)

        self.client.login(username='delegate', password='password')
        detail_response = self.client.get(reverse('document_detail', args=[document.pk]))
        response = self.client.post(
            reverse('document_update_movement', args=[document.pk]),
            {
                'new_status': 'closed',
                'to_department': self.ops_department.pk,
                'remarks': 'Closed by delegated receiver.',
            },
        )

        self.assertContains(detail_response, 'Delegation Information')
        self.assertRedirects(response, reverse('document_detail', args=[document.pk]))
        document.refresh_from_db()
        movement = DocumentMovement.objects.filter(document=document, action='closed').get()
        self.assertEqual(document.status, 'closed')
        self.assertEqual(movement.performed_by, self.delegated_receiver)

    def test_unrelated_user_cannot_view_or_close_delegated_document(self):
        document = self.create_document(
            designated_person=self.designated_person,
            status='in_progress',
        )
        self.create_delegation(document)

        self.client.login(username='unrelated', password='password')
        detail_response = self.client.get(reverse('document_detail', args=[document.pk]))
        close_response = self.client.post(
            reverse('document_update_movement', args=[document.pk]),
            {
                'new_status': 'closed',
                'remarks': 'Try closing.',
            },
        )

        self.assertEqual(detail_response.status_code, 403)
        self.assertEqual(close_response.status_code, 403)
        document.refresh_from_db()
        self.assertEqual(document.status, 'in_progress')

    def test_bulk_forward_requires_confirmation_and_updates_each_document(self):
        first_document = self.create_document(
            subject='First bulk forward',
            designated_person=self.designated_person,
            destination_department=self.ops_department,
        )
        second_document = self.create_document(
            subject='Second bulk forward',
            designated_person=self.designated_person,
            destination_department=self.ops_department,
        )

        self.client.login(username='polu', password='password')
        confirm_response = self.client.post(
            reverse('bulk_forward_documents'),
            {
                'selected_document_ids': [first_document.pk, second_document.pk],
                'next': reverse('pending_documents'),
            },
        )

        self.assertEqual(confirm_response.status_code, 200)
        self.assertContains(confirm_response, 'Confirm Bulk Forward')
        first_document.refresh_from_db()
        self.assertEqual(first_document.status, 'created')

        response = self.client.post(
            reverse('bulk_forward_documents'),
            {
                'confirm': '1',
                'selected_document_ids': [first_document.pk, second_document.pk],
                'remarks': 'Bulk test forward',
                'next': reverse('pending_documents'),
            },
        )

        self.assertRedirects(response, reverse('pending_documents'))
        first_document.refresh_from_db()
        second_document.refresh_from_db()

        self.assertEqual(first_document.status, 'forwarded')
        self.assertEqual(second_document.status, 'forwarded')
        self.assertEqual(first_document.current_department, self.ops_department)
        self.assertEqual(second_document.current_department, self.ops_department)
        self.assertEqual(
            DocumentMovement.objects.filter(action='forwarded').count(),
            2,
        )
        self.assertEqual(len(mail.outbox), 2)

    def test_bulk_forward_empty_selection_shows_message(self):
        self.client.login(username='polu', password='password')

        response = self.client.post(
            reverse('bulk_forward_documents'),
            {'next': reverse('pending_documents')},
            follow=True,
        )

        self.assertContains(response, 'Please select at least one document to forward.')

    def test_management_cannot_access_bulk_actions(self):
        document = self.create_document()
        self.client.login(username='manager', password='password')

        forward_response = self.client.post(
            reverse('bulk_forward_documents'),
            {'selected_document_ids': [document.pk]},
        )
        close_response = self.client.post(
            reverse('bulk_close_documents'),
            {'selected_document_ids': [document.pk]},
        )

        self.assertEqual(forward_response.status_code, 403)
        self.assertEqual(close_response.status_code, 403)

    def test_assigned_user_bulk_closes_only_assigned_documents(self):
        own_document = self.create_document(
            subject='Assigned close',
            designated_person=self.designated_person,
            status='in_progress',
        )
        other_document = self.create_document(
            subject='Other close',
            status='in_progress',
        )

        self.client.login(username='recipient', password='password')
        confirm_response = self.client.post(
            reverse('bulk_close_documents'),
            {
                'selected_document_ids': [own_document.pk, other_document.pk],
                'next': reverse('my_assigned_documents'),
            },
        )

        self.assertContains(confirm_response, 'Confirm Bulk Close')
        self.assertContains(confirm_response, 'No permission')
        own_document.refresh_from_db()
        self.assertEqual(own_document.status, 'in_progress')

        response = self.client.post(
            reverse('bulk_close_documents'),
            {
                'confirm': '1',
                'selected_document_ids': [own_document.pk],
                'skipped_count': '1',
                'remarks': 'Finished assigned work.',
                'next': reverse('my_assigned_documents'),
            },
            follow=True,
        )

        own_document.refresh_from_db()
        other_document.refresh_from_db()
        movement = DocumentMovement.objects.get(document=own_document)

        self.assertEqual(own_document.status, 'closed')
        self.assertEqual(other_document.status, 'in_progress')
        self.assertEqual(movement.action, 'closed')
        self.assertEqual(movement.performed_by, self.designated_person)
        self.assertContains(
            response,
            '1 documents closed successfully. 1 skipped due to permission/status.'
        )

    def test_delegated_user_can_bulk_close_delegated_document(self):
        document = self.create_document(
            subject='Delegated bulk close',
            designated_person=self.designated_person,
            status='in_progress',
        )
        self.create_delegation(document)

        self.client.login(username='delegate', password='password')
        response = self.client.post(
            reverse('bulk_close_documents'),
            {
                'confirm': '1',
                'selected_document_ids': [document.pk],
                'remarks': 'Finished delegated work.',
                'next': reverse('my_assigned_documents'),
            },
        )

        self.assertRedirects(response, reverse('my_assigned_documents'))
        document.refresh_from_db()
        movement = DocumentMovement.objects.get(document=document, action='closed')
        self.assertEqual(document.status, 'closed')
        self.assertEqual(movement.performed_by, self.delegated_receiver)

    def test_receiving_desk_can_bulk_close_non_closed_documents(self):
        document = self.create_document(status='in_progress')

        self.client.login(username='polu', password='password')
        response = self.client.post(
            reverse('bulk_close_documents'),
            {
                'confirm': '1',
                'selected_document_ids': [document.pk],
                'remarks': 'Closed by receiving desk.',
                'next': reverse('pending_documents'),
            },
        )

        self.assertRedirects(response, reverse('pending_documents'))
        document.refresh_from_db()
        self.assertEqual(document.status, 'closed')
        self.assertEqual(DocumentMovement.objects.get(document=document).action, 'closed')

    def test_my_assigned_documents_shows_only_open_documents_for_current_user(self):
        own_open = self.create_document(
            subject='Assigned open',
            designated_person=self.designated_person,
            status='in_progress',
        )
        own_closed = self.create_document(
            subject='Assigned closed',
            designated_person=self.designated_person,
            status='closed',
        )
        other_open = self.create_document(
            subject='Other open',
            status='in_progress',
        )
        delegated_open = self.create_document(
            subject='Delegated open',
            designated_person=self.first_boss,
            status='in_progress',
        )
        delegated_closed = self.create_document(
            subject='Delegated closed',
            designated_person=self.first_boss,
            status='closed',
        )
        self.create_delegation(
            delegated_open,
            original_recipient=self.first_boss,
            delegated_recipient=self.designated_person,
            delegated_by=self.first_boss,
        )
        self.create_delegation(
            delegated_closed,
            original_recipient=self.first_boss,
            delegated_recipient=self.designated_person,
            delegated_by=self.first_boss,
        )

        self.client.login(username='recipient', password='password')
        response = self.client.get(reverse('my_assigned_documents'))

        self.assertContains(response, own_open.tracking_id)
        self.assertNotContains(response, own_closed.tracking_id)
        self.assertNotContains(response, other_open.tracking_id)
        self.assertContains(response, delegated_open.tracking_id)
        self.assertNotContains(response, delegated_closed.tracking_id)
        self.assertContains(response, 'Close Selected')

    def test_smtp_failure_is_saved_without_crashing_forward_update(self):
        document = self.create_document(designated_person=self.designated_person)

        self.client.login(username='polu', password='password')
        with mock.patch('documents.services.EmailMessage.send', side_effect=Exception('SMTP down')):
            response = self.post_forward_update(document)

        self.assertRedirects(response, reverse('document_detail', args=[document.pk]))
        document.refresh_from_db()
        self.assertEqual(document.status, 'forwarded')
        self.assertFalse(document.notification_sent)
        self.assertEqual(document.notification_error, 'SMTP down')

    def test_document_search_includes_reference_and_user_email_fields(self):
        courier_rate = CourierRate.objects.create(
            description='Courier Search Packet',
            quantity='0-500 grams',
            amount=Decimal('25.00'),
        )
        document = self.create_document(
            entry_type='outward',
            reference_no='MEMO-SEARCH-42',
            addressed_to_name='Recipient Person',
            designated_person=self.designated_person,
            source_division=self.division,
            source_region=self.region,
            source_area=self.area,
            source_branch=self.branch,
            courier_rate=courier_rate,
            outward_description='Courier Search Packet',
            outward_quantity='0-500 grams',
        )

        self.client.login(username='manager', password='password')
        reference_response = self.client.get(reverse('document_list'), {'q': 'MEMO-SEARCH-42'})
        email_response = self.client.get(reverse('document_list'), {'q': 'recipient@example.org'})
        branch_response = self.client.get(reverse('document_list'), {'q': 'Banasree Branch'})
        courier_response = self.client.get(reverse('document_list'), {'q': 'Courier Search Packet'})
        quantity_response = self.client.get(reverse('document_list'), {'q': '0-500 grams'})

        self.assertContains(reference_response, document.tracking_id)
        self.assertContains(email_response, document.tracking_id)
        self.assertContains(branch_response, document.tracking_id)
        self.assertContains(courier_response, document.tracking_id)
        self.assertContains(quantity_response, document.tracking_id)

    def test_dashboard_counts_split_pending_forwarded_and_closed_documents(self):
        self.create_document(subject='Pending normal', status='created')
        self.create_document(subject='Pending urgent', status='in_progress', priority='urgent')
        self.create_document(subject='Forwarded urgent', status='forwarded', priority='urgent')
        self.create_document(subject='Closed urgent', status='closed', priority='urgent')

        self.client.login(username='polu', password='password')
        response = self.client.get(reverse('dashboard'))

        self.assertEqual(response.context['total_documents'], 4)
        self.assertEqual(response.context['pending_documents'], 2)
        self.assertEqual(response.context['forwarded_documents'], 1)
        self.assertEqual(response.context['closed_documents'], 1)
        self.assertEqual(response.context['urgent_documents'], 1)
        self.assertContains(response, 'Forwarded Documents')
        self.assertNotContains(response, 'My Department Pending')

    def test_assigned_user_dashboard_shows_only_their_assigned_documents(self):
        own_open = self.create_document(
            subject='My open urgent',
            designated_person=self.designated_person,
            status='in_progress',
            priority='urgent',
        )
        own_forwarded = self.create_document(
            subject='My forwarded',
            designated_person=self.designated_person,
            status='forwarded',
            priority='urgent',
        )
        own_closed = self.create_document(
            subject='My closed',
            designated_person=self.designated_person,
            status='closed',
            priority='urgent',
        )
        other_document = self.create_document(
            subject='Someone else recent',
            status='created',
            priority='urgent',
        )

        self.client.login(username='recipient', password='password')
        response = self.client.get(reverse('dashboard'))

        self.assertFalse(response.context['can_view_system_dashboard'])
        self.assertEqual(response.context['dashboard_table_title'], 'My Assigned Documents')
        self.assertEqual(response.context['total_documents'], 3)
        self.assertNotIn('pending_documents', response.context)
        self.assertNotIn('forwarded_documents', response.context)
        self.assertEqual(response.context['closed_documents'], 1)
        self.assertEqual(response.context['urgent_pending'], 2)
        self.assertEqual(response.context['my_assigned_documents_count'], 2)
        self.assertContains(response, own_open.tracking_id)
        self.assertContains(response, own_forwarded.tracking_id)
        self.assertContains(response, own_closed.tracking_id)
        self.assertNotContains(response, other_document.tracking_id)
        self.assertContains(response, 'Total Assigned to Me')
        self.assertContains(response, 'Closed by Me')
        self.assertContains(response, 'Urgent Assigned')
        self.assertContains(response, 'My Assigned Documents')
        self.assertNotContains(response, 'Recent Documents')
        self.assertNotContains(response, 'Pending Documents')
        self.assertNotContains(response, 'Forwarded Documents')

    def test_assigned_user_dashboard_counts_delegations(self):
        delegated_to_me = self.create_document(
            subject='Delegated to recipient',
            designated_person=self.first_boss,
            status='in_progress',
        )
        delegated_by_me = self.create_document(
            subject='Delegated by recipient',
            designated_person=self.designated_person,
            status='in_progress',
        )
        self.create_delegation(
            delegated_to_me,
            original_recipient=self.first_boss,
            delegated_recipient=self.designated_person,
            delegated_by=self.first_boss,
        )
        self.create_delegation(delegated_by_me)

        self.client.login(username='recipient', password='password')
        response = self.client.get(reverse('dashboard'))

        self.assertEqual(response.context['delegated_to_me_count'], 1)
        self.assertEqual(response.context['delegated_by_me_count'], 1)
        self.assertContains(response, 'Delegated to Me')
        self.assertContains(response, 'Delegated by Me')

    def test_assigned_user_direct_document_views_are_limited_to_assigned_documents(self):
        own_pending = self.create_document(
            subject='Own pending',
            designated_person=self.designated_person,
            status='in_progress',
        )
        own_forwarded = self.create_document(
            subject='Own forwarded',
            designated_person=self.designated_person,
            status='forwarded',
        )
        other_pending = self.create_document(subject='Other pending', status='in_progress')
        other_forwarded = self.create_document(subject='Other forwarded', status='forwarded')

        self.client.login(username='recipient', password='password')

        list_response = self.client.get(reverse('document_list'))
        pending_response = self.client.get(reverse('pending_documents'))
        forwarded_response = self.client.get(reverse('forwarded_documents'))
        own_detail_response = self.client.get(reverse('document_detail', args=[own_pending.pk]))
        other_detail_response = self.client.get(reverse('document_detail', args=[other_pending.pk]))

        self.assertContains(list_response, own_pending.tracking_id)
        self.assertContains(list_response, own_forwarded.tracking_id)
        self.assertNotContains(list_response, other_pending.tracking_id)
        self.assertNotContains(list_response, other_forwarded.tracking_id)

        self.assertContains(pending_response, own_pending.tracking_id)
        self.assertNotContains(pending_response, own_forwarded.tracking_id)
        self.assertNotContains(pending_response, other_pending.tracking_id)

        self.assertContains(forwarded_response, own_forwarded.tracking_id)
        self.assertNotContains(forwarded_response, own_pending.tracking_id)
        self.assertNotContains(forwarded_response, other_forwarded.tracking_id)

        self.assertEqual(own_detail_response.status_code, 200)
        self.assertEqual(other_detail_response.status_code, 403)

    def test_pending_documents_excludes_forwarded_and_closed_documents(self):
        pending_document = self.create_document(subject='Visible pending', status='created')
        forwarded_document = self.create_document(subject='Hidden forwarded', status='forwarded')
        closed_document = self.create_document(subject='Hidden closed', status='closed')

        self.client.login(username='polu', password='password')
        response = self.client.get(reverse('pending_documents'))

        self.assertContains(response, pending_document.tracking_id)
        self.assertNotContains(response, forwarded_document.tracking_id)
        self.assertNotContains(response, closed_document.tracking_id)

    def test_forwarded_documents_page_shows_only_forwarded_documents(self):
        pending_document = self.create_document(subject='Not forwarded', status='created')
        forwarded_document = self.create_document(subject='Visible forwarded', status='forwarded')
        closed_document = self.create_document(subject='Closed document', status='closed')

        self.client.login(username='polu', password='password')
        response = self.client.get(reverse('forwarded_documents'))

        self.assertContains(response, 'Forwarded Documents')
        self.assertContains(response, forwarded_document.tracking_id)
        self.assertNotContains(response, pending_document.tracking_id)
        self.assertNotContains(response, closed_document.tracking_id)

    def test_reports_dashboard_matches_forwarded_and_hierarchy_workflow(self):
        pending_document = self.create_document(
            subject='Pending report item',
            status='in_progress',
            source_division=self.division,
        )
        forwarded_document = self.create_document(
            subject='Forwarded report item',
            status='forwarded',
            priority='urgent',
        )
        closed_document = self.create_document(
            subject='Closed report item',
            status='closed',
            priority='urgent',
        )
        DocumentMovement.objects.create(
            document=forwarded_document,
            action='forwarded',
            performed_by=self.receiving_user,
        )

        self.client.login(username='manager', password='password')
        response = self.client.get(reverse('reports_dashboard'))

        self.assertEqual(response.context['total_documents'], 3)
        self.assertEqual(response.context['pending_documents'], 1)
        self.assertEqual(response.context['forwarded_documents'], 1)
        self.assertEqual(response.context['closed_documents'], 1)
        self.assertEqual(response.context['urgent_open'], 1)
        self.assertEqual(response.context['forwarded_today'], 1)
        self.assertEqual(response.context['closed_today'], 1)

        aging_documents = list(response.context['aging_documents'])
        self.assertIn(pending_document, aging_documents)
        self.assertIn(forwarded_document, aging_documents)
        self.assertNotIn(closed_document, aging_documents)

        status_counts = {
            item['label']: item['count']
            for item in response.context['status_summary']
        }
        entry_type_counts = {
            item['label']: item['count']
            for item in response.context['entry_type_summary']
        }
        division_counts = {
            item['name']: item['document_count']
            for item in response.context['division_documents']
        }

        self.assertEqual(status_counts['Forwarded'], 1)
        self.assertEqual(status_counts['Returned'], 0)
        self.assertEqual(entry_type_counts['Inward'], 3)
        self.assertEqual(entry_type_counts['Outward'], 0)
        self.assertEqual(division_counts[self.division.name], 1)
        self.assertEqual(division_counts['External / Not Specified'], 2)
        self.assertContains(response, 'Forwarded Documents')
        self.assertContains(response, 'Urgent Open')
        self.assertContains(response, 'Forwarded Today')
        self.assertContains(response, 'Closed Today')
        self.assertContains(response, 'Division-wise Documents')
        self.assertContains(response, 'Aging Open Documents')
        self.assertContains(response, 'No outwarding cost records found.')
        self.assertNotContains(response, 'Zone-wise Documents')
        self.assertNotContains(response, 'Aging Pending Documents')

    def test_delegation_report_permission_and_content(self):
        document = self.create_document(
            subject='Delegated report item',
            designated_person=self.designated_person,
            status='in_progress',
        )
        self.create_delegation(
            document,
            reason='Training leave.',
            delegated_by=self.designated_person,
        )

        self.client.login(username='manager', password='password')
        manager_response = self.client.get(reverse('delegation_report'))
        self.client.logout()
        self.client.login(username='unrelated', password='password')
        unrelated_response = self.client.get(reverse('delegation_report'))

        self.assertEqual(manager_response.status_code, 200)
        self.assertContains(manager_response, document.tracking_id)
        self.assertContains(manager_response, 'Training leave.')
        self.assertEqual(unrelated_response.status_code, 403)

    def test_source_display_uses_most_specific_hierarchy_value(self):
        document = self.create_document(
            source_division=self.division,
            source_region=self.region,
            source_area=self.area,
            source_branch=self.branch,
        )

        self.assertEqual(document.source_display, self.branch)

        document.source_branch = None
        self.assertEqual(document.source_display, self.area)

    def test_source_display_uses_external_source_information(self):
        pksf_document = self.create_document(
            source_type='external',
            external_organization_type='pksf',
        )
        bank_document = self.create_document(
            source_type='external',
            external_organization_type='bank',
            external_organization_name='City Bank',
            external_branch_name='Gulshan',
        )
        government_document = self.create_document(
            source_type='external',
            external_organization_type='government_offices',
            external_organization_name='Ministry of Finance',
            external_branch_name='Dhaka',
        )

        self.assertEqual(pksf_document.source_display, 'PKSF')
        self.assertEqual(bank_document.source_display, 'City Bank - Gulshan')
        self.assertEqual(government_document.source_display, 'Ministry of Finance - Dhaka')

    def test_document_detail_displays_government_office_details(self):
        document = self.create_document(
            source_type='external',
            external_organization_type='government_offices',
            external_organization_name='Ministry of Finance',
            external_branch_name='Dhaka',
        )

        self.client.login(username='polu', password='password')
        response = self.client.get(reverse('document_detail', args=[document.pk]))

        self.assertContains(response, 'Government Offices')
        self.assertContains(response, 'Ministry of Finance')
        self.assertContains(response, 'Dhaka')

    def test_document_search_includes_external_source_fields(self):
        matching_document = self.create_document(
            source_type='external',
            external_organization_type='bank',
            external_organization_name='Eastern Finance',
            external_branch_name='Motijheel',
        )
        other_document = self.create_document(subject='Other document')

        self.client.login(username='polu', password='password')
        response = self.client.get(reverse('document_list'), {'q': 'Motijheel'})

        self.assertContains(response, matching_document.tracking_id)
        self.assertNotContains(response, other_document.tracking_id)

    def test_hierarchy_api_returns_active_children(self):
        inactive_region = Region.objects.create(
            division=self.division,
            name='Inactive Region',
            is_active=False,
        )
        self.client.login(username='polu', password='password')

        response = self.client.get(reverse('get_regions'), {'division_id': self.division.pk})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            json.loads(response.content),
            [{'id': self.region.pk, 'name': self.region.name}],
        )
        self.assertNotContains(response, inactive_region.name)

    def test_hierarchy_api_returns_areas_and_branches(self):
        self.client.login(username='polu', password='password')

        area_response = self.client.get(reverse('get_areas'), {'region_id': self.region.pk})
        branch_response = self.client.get(reverse('get_branches'), {'area_id': self.area.pk})

        self.assertEqual(
            json.loads(area_response.content),
            [{'id': self.area.pk, 'name': self.area.name}],
        )
        self.assertEqual(
            json.loads(branch_response.content),
            [{'id': self.branch.pk, 'name': self.branch.name}],
        )

    def test_courier_rate_api_returns_active_rate_only(self):
        active_rate = CourierRate.objects.create(
            description='ছোট পলি প্যাকেট',
            quantity='কন্টিনেন্টাল এর লোগো সম্বলিত পলি',
            amount=Decimal('95.00'),
            remarks='Standard packet',
        )
        inactive_rate = CourierRate.objects.create(
            description='Inactive service',
            quantity='1',
            amount=Decimal('1.00'),
            is_active=False,
        )

        self.client.login(username='polu', password='password')
        active_response = self.client.get(
            reverse('get_courier_rate'),
            {'courier_rate_id': active_rate.pk},
        )
        inactive_response = self.client.get(
            reverse('get_courier_rate'),
            {'courier_rate_id': inactive_rate.pk},
        )

        self.assertEqual(active_response.status_code, 200)
        self.assertEqual(
            json.loads(active_response.content),
            {
                'id': active_rate.pk,
                'description': active_rate.description,
                'quantity': active_rate.quantity,
                'amount': '95.00',
                'remarks': 'Standard packet',
            },
        )
        self.assertEqual(inactive_response.status_code, 404)

    def test_excel_export_includes_notification_columns(self):
        self.create_document(
            reference_no='REF-XLS',
            addressed_to_name='Recipient Person',
            addressed_to_designation='Director',
            designated_person=self.designated_person,
            first_boss=self.first_boss,
            notification_error='SMTP down',
            source_division=self.division,
            source_region=self.region,
            source_area=self.area,
            source_branch=self.branch,
        )

        self.client.login(username='manager', password='password')
        response = self.client.get(reverse('export_documents_excel'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content))
        headers = [cell.value for cell in workbook.active[1]]

        self.assertIn('Reference No', headers)
        self.assertIn('Addressed To Name', headers)
        self.assertIn('Designated Person Email', headers)
        self.assertIn('First Boss Email', headers)
        self.assertIn('Notification Error', headers)
        self.assertIn('Source Type', headers)
        self.assertIn('External Organization Type', headers)
        self.assertIn('External Organization Name', headers)
        self.assertIn('External Branch', headers)
        self.assertIn('Source Division', headers)
        self.assertIn('Source Region', headers)
        self.assertIn('Source Area', headers)
        self.assertIn('Source Branch', headers)
        self.assertIn('Courier Description', headers)
        self.assertIn('Courier Quantity', headers)
        self.assertIn('Courier Amount', headers)
        self.assertIn('Courier Rate Source', headers)


class UserImportCommandTests(TestCase):
    def create_workbook(self, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Ready_Import'
        worksheet.append([
            'username',
            'email',
            'first_name',
            'last_name',
            'department',
            'role',
            'designation',
            'phone',
            'group',
            'is_active',
        ])

        for row in rows:
            worksheet.append(row)

        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        workbook.save(temp_file.name)
        return temp_file.name

    def run_import(self, workbook_path, *args):
        try:
            call_command(
                'import_users_from_excel',
                workbook_path,
                '--sheet',
                'Ready_Import',
                '--default-password',
                'ChangeThis@123',
                *args,
            )
        finally:
            os.unlink(workbook_path)

    def test_import_uses_staff_id_as_username(self):
        workbook_path = self.create_workbook([
            [
                '0070521009',
                'sums@padakhep.org',
                'Md. Saleh Bin',
                'Sums',
                'Executive Director Secretariat',
                '',
                'Executive Director',
                '01730024515',
                '',
                'TRUE',
            ],
        ])

        self.run_import(workbook_path)

        user = get_user_model().objects.get(username='0070521009')
        self.assertEqual(user.email, 'sums@padakhep.org')
        self.assertTrue(user.check_password('ChangeThis@123'))
        self.assertEqual(user.profile.department.name, 'Executive Director Secretariat')
        self.assertEqual(user.profile.designation, 'Executive Director')

    def test_replace_name_usernames_recreates_users_and_restores_document_links(self):
        user_model = get_user_model()
        department = Department.objects.create(name='HR', code='HR')
        old_user = user_model.objects.create_user(
            username='nusrat',
            password='password',
            email='nusrat@padakhep.org',
        )
        UserProfile.objects.create(
            user=old_user,
            department=department,
            designation='Officer',
        )
        document = Document.objects.create(
            entry_type='inward',
            subject='Assigned memo',
            designated_person=old_user,
            first_boss=old_user,
            created_by=old_user,
        )
        movement = DocumentMovement.objects.create(
            document=document,
            action='status_updated',
            performed_by=old_user,
            remarks='Old user movement.',
        )
        protected_user = user_model.objects.create_user(
            username='editor.name',
            password='password',
            email='editor@padakhep.org',
        )
        protected_group, _ = Group.objects.get_or_create(name='Receiving Desk')
        protected_user.groups.add(protected_group)
        UserProfile.objects.create(user=protected_user, department=department)

        workbook_path = self.create_workbook([
            [
                '0000000001',
                'nusrat@padakhep.org',
                'Nusrat',
                'Jahan',
                'HR',
                '',
                'Officer',
                '01700000001',
                '',
                'TRUE',
            ],
            [
                '0000000002',
                'editor@padakhep.org',
                'Editor',
                'Person',
                'HR',
                '',
                'Editor',
                '01700000002',
                '',
                'TRUE',
            ],
        ])

        self.run_import(workbook_path, '--replace-name-usernames')

        self.assertFalse(user_model.objects.filter(username='nusrat').exists())
        new_user = user_model.objects.get(username='0000000001')
        document.refresh_from_db()
        movement.refresh_from_db()
        protected_user.refresh_from_db()

        self.assertEqual(document.designated_person, new_user)
        self.assertEqual(document.first_boss, new_user)
        self.assertEqual(document.created_by, new_user)
        self.assertEqual(movement.performed_by, new_user)
        self.assertEqual(protected_user.username, 'editor.name')
        self.assertFalse(user_model.objects.filter(username='0000000002').exists())
