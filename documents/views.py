from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Case, Count, IntegerField, Q, Sum, Value, When
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .forms import (
    DocumentDelegationForm,
    DocumentForm,
    DocumentMovementForm,
    UserDelegationRuleForm,
)
from .models import (
    Area,
    Branch,
    CourierRate,
    Department,
    Document,
    DocumentDelegation,
    DocumentMovement,
    Region,
    UserDelegationRule,
    Zone,
)
from .services import send_forward_notification


EDITOR_GROUPS = ['Admin', 'Receiving Desk', 'HR', 'ADT', 'OPS']
REPORT_GROUPS = ['Admin', 'HR', 'ADT', 'OPS', 'Management']
BULK_FORWARD_GROUPS = ['Admin', 'Receiving Desk']
BULK_CLOSE_ALL_GROUPS = ['Admin', 'Receiving Desk']
VIEW_ONLY_GROUPS = ['Management']
SYSTEM_VIEW_GROUPS = EDITOR_GROUPS + VIEW_ONLY_GROUPS


def user_in_groups(user, group_names):
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=group_names).exists()


def user_has_named_group(user, group_names):
    if not user.is_authenticated:
        return False
    return user.groups.filter(name__in=group_names).exists()


def can_edit_documents(user):
    return user_in_groups(user, EDITOR_GROUPS)


def can_view_system_documents(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return user_has_named_group(user, SYSTEM_VIEW_GROUPS)


def is_designated_person_for_document(user, document):
    return (
        user.is_authenticated and
        document.designated_person_id == user.id
    )


def active_delegation_for_document(document):
    prefetched_delegations = getattr(document, '_prefetched_objects_cache', {}).get('delegations')
    if prefetched_delegations is not None:
        return next((delegation for delegation in prefetched_delegations if delegation.is_active), None)
    return document.delegations.filter(is_active=True).select_related(
        'original_recipient',
        'delegated_recipient',
        'delegated_by',
    ).first()


def is_active_delegated_recipient_for_document(user, document):
    if not user.is_authenticated:
        return False
    delegation = active_delegation_for_document(document)
    return bool(delegation and delegation.delegated_recipient_id == user.id)


def can_delegate_document(user, document):
    return (
        is_designated_person_for_document(user, document) and
        document.status != 'closed' and
        not active_delegation_for_document(document)
    )


def can_bulk_forward(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return user_has_named_group(user, BULK_FORWARD_GROUPS)


def can_bulk_close(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    if user_has_named_group(user, BULK_CLOSE_ALL_GROUPS):
        return True
    if user_has_named_group(user, VIEW_ONLY_GROUPS):
        return False
    return True


def can_user_close_document(user, document):
    if document.status == 'closed':
        return False
    if not user.is_authenticated:
        return False
    if user.is_superuser or user_has_named_group(user, BULK_CLOSE_ALL_GROUPS):
        return True
    return (
        document.designated_person_id == user.id or
        is_active_delegated_recipient_for_document(user, document)
    )


def can_view_document(user, document):
    return (
        can_view_system_documents(user) or
        is_designated_person_for_document(user, document) or
        is_active_delegated_recipient_for_document(user, document)
    )


def can_update_document_status(user, document):
    return (
        can_edit_documents(user) or
        is_designated_person_for_document(user, document) or
        is_active_delegated_recipient_for_document(user, document)
    )


def can_view_reports(user):
    return user_in_groups(user, REPORT_GROUPS)


def can_view_delegation_reports(user):
    return user_in_groups(user, ['Admin', 'Receiving Desk', 'Management'])


def assigned_documents_filter(user):
    return (
        Q(designated_person=user) |
        Q(delegations__is_active=True, delegations__delegated_recipient=user)
    )


def active_temporary_delegation_rule_for_user(user, today=None):
    if not user or not user.is_authenticated:
        return None
    if today is None:
        today = timezone.now().date()
    return UserDelegationRule.objects.filter(
        user=user,
        is_active=True,
        start_date__lte=today,
        end_date__gte=today,
        backup_receiver__is_active=True,
    ).select_related('backup_receiver').order_by('-created_at').first()


def movement_action_from_status(status):
    return {
        'forwarded': 'forwarded',
        'physical_received': 'physical_received',
        'received_by_department': 'received',
        'closed': 'closed',
        'returned': 'returned',
    }.get(status, 'status_updated')


def courier_rate_source(document):
    if document.outward_is_manual:
        return 'Manual / Other'
    if document.courier_rate_id:
        return 'Predefined Rate'
    return ''


def receiving_desk_department():
    return (
        Department.objects.filter(name='Receiving Desk', is_active=True).first()
        or Department.objects.filter(name='Receiving Desk').first()
    )


def user_profile_department(user):
    profile = getattr(user, 'profile', None)
    if profile and profile.department:
        return profile.department
    return None


def document_form_department_context(request):
    receiving_department = receiving_desk_department()
    user_department = user_profile_department(request.user)

    return {
        'receiving_desk_department_id': receiving_department.pk if receiving_department else '',
        'user_department_id': user_department.pk if user_department else '',
    }


def active_hierarchy_options(queryset):
    return JsonResponse(list(queryset.values('id', 'name')), safe=False)


@login_required
def get_regions(request):
    division_id = request.GET.get('division_id', '').strip()
    if not division_id.isdigit():
        return JsonResponse([], safe=False)

    regions = Region.objects.filter(
        division_id=division_id,
        is_active=True,
    ).order_by('name')
    return active_hierarchy_options(regions)


@login_required
def get_areas(request):
    region_id = request.GET.get('region_id', '').strip()
    if not region_id.isdigit():
        return JsonResponse([], safe=False)

    areas = Area.objects.filter(
        region_id=region_id,
        is_active=True,
    ).order_by('name')
    return active_hierarchy_options(areas)


@login_required
def get_branches(request):
    area_id = request.GET.get('area_id', '').strip()
    if not area_id.isdigit():
        return JsonResponse([], safe=False)

    branches = Branch.objects.filter(
        area_id=area_id,
        is_active=True,
    ).order_by('name')
    return active_hierarchy_options(branches)


@login_required
def get_courier_rate(request):
    courier_rate_id = request.GET.get('courier_rate_id', '').strip()
    if not courier_rate_id.isdigit():
        return JsonResponse(
            {'error': 'Courier rate not found.'},
            status=404,
        )

    courier_rate = CourierRate.objects.filter(
        pk=courier_rate_id,
        is_active=True,
    ).first()

    if not courier_rate:
        return JsonResponse(
            {'error': 'Courier rate not found.'},
            status=404,
        )

    return JsonResponse({
        'id': courier_rate.id,
        'description': courier_rate.description,
        'quantity': courier_rate.quantity,
        'amount': format(courier_rate.amount, '.2f'),
        'remarks': courier_rate.remarks,
    })


@login_required
def get_user_info(request):
    user_id = request.GET.get('user_id', '').strip()
    if not user_id.isdigit():
        return JsonResponse(
            {'error': 'User not found.'},
            status=404,
        )

    user = get_user_model().objects.filter(
        pk=user_id,
        is_active=True,
    ).select_related('profile__department').first()

    if not user:
        return JsonResponse(
            {'error': 'User not found.'},
            status=404,
        )

    profile = getattr(user, 'profile', None)
    department = profile.department if profile and profile.department else None

    return JsonResponse({
        'id': user.id,
        'name': user.get_full_name() or user.username,
        'designation': profile.designation if profile else '',
        'department': department.name if department else '',
    })


def permission_denied_page(request, message='You do not have permission to perform this action.'):
    return render(
        request,
        'documents/permission_denied.html',
        {'message': message},
        status=403
    )


def safe_next_url(request, default_name='document_list'):
    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url
    return reverse(default_name)


def selected_document_ids(request):
    ids = []
    for raw_id in request.POST.getlist('selected_document_ids'):
        raw_id = raw_id.strip()
        if raw_id.isdigit():
            ids.append(int(raw_id))
    return list(dict.fromkeys(ids))


def posted_int(request, field_name, default=0):
    value = request.POST.get(field_name, '')
    if str(value).isdigit():
        return int(value)
    return default


def selected_documents_queryset(document_ids):
    return Document.objects.filter(pk__in=document_ids).select_related(
        'source_department',
        'destination_department',
        'current_department',
        'designated_person',
        'first_boss',
    ).prefetch_related(
        'delegations',
    ).order_by('tracking_id')


def skipped_item(document=None, reason=''):
    return {
        'document': document,
        'reason': reason,
    }


def partition_bulk_forward_documents(documents, requested_count):
    actionable = []
    skipped = []

    for document in documents:
        if document.status == 'closed':
            skipped.append(skipped_item(document, 'Already closed'))
        else:
            actionable.append(document)

    missing_count = requested_count - len(documents)
    for _ in range(max(missing_count, 0)):
        skipped.append(skipped_item(reason='Document was not found'))

    return actionable, skipped


def partition_bulk_close_documents(user, documents, requested_count):
    actionable = []
    skipped = []

    for document in documents:
        if can_user_close_document(user, document):
            actionable.append(document)
        elif document.status == 'closed':
            skipped.append(skipped_item(document, 'Already closed'))
        else:
            skipped.append(skipped_item(document, 'No permission'))

    missing_count = requested_count - len(documents)
    for _ in range(max(missing_count, 0)):
        skipped.append(skipped_item(reason='Document was not found'))

    return actionable, skipped


def render_bulk_confirmation(
    request,
    action,
    documents,
    skipped,
    remarks_error='',
):
    is_forward = action == 'forward'
    return render(request, 'documents/bulk_action_confirm.html', {
        'action': action,
        'page_title': 'Confirm Bulk Forward' if is_forward else 'Confirm Bulk Close',
        'documents': documents,
        'skipped_items': skipped,
        'selected_document_ids': [document.pk for document in documents],
        'bulk_url_name': 'bulk_forward_documents' if is_forward else 'bulk_close_documents',
        'remarks_label': 'Forwarding Remarks' if is_forward else 'Closing Remarks',
        'remarks_required': not is_forward,
        'remarks_default': request.POST.get('remarks', ''),
        'remarks_error': remarks_error,
        'skipped_count': len(skipped) + posted_int(request, 'skipped_count'),
        'next_url': safe_next_url(
            request,
            'document_list' if is_forward else 'my_assigned_documents',
        ),
    })


def filter_documents_from_request(request, base_queryset=None):
    if base_queryset is None:
        documents = Document.objects.select_related(
            'source_zone',
            'source_division',
            'source_region',
            'source_area',
            'source_branch',
            'source_department',
            'destination_department',
            'current_department',
            'created_by',
            'designated_person',
            'first_boss',
            'courier_rate',
        ).order_by('-created_at')
    else:
        documents = base_queryset

    query = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()
    priority = request.GET.get('priority', '').strip()
    department_id = request.GET.get('department', '').strip()
    zone_id = request.GET.get('zone', '').strip()
    entry_type = request.GET.get('entry_type', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    if query:
        documents = documents.filter(
            Q(tracking_id__icontains=query) |
            Q(reference_no__icontains=query) |
            Q(subject__icontains=query) |
            Q(document_type__icontains=query) |
            Q(addressed_to_name__icontains=query) |
            Q(addressed_to_designation__icontains=query) |
            Q(designated_person__username__icontains=query) |
            Q(designated_person__email__icontains=query) |
            Q(first_boss__username__icontains=query) |
            Q(first_boss__email__icontains=query) |
            Q(source_type__icontains=query) |
            Q(external_organization_type__icontains=query) |
            Q(external_organization_name__icontains=query) |
            Q(external_branch_name__icontains=query) |
            Q(outward_description__icontains=query) |
            Q(outward_quantity__icontains=query) |
            Q(courier_rate__description__icontains=query) |
            Q(courier_rate__quantity__icontains=query) |
            Q(remarks__icontains=query) |
            Q(source_division__name__icontains=query) |
            Q(source_division__code__icontains=query) |
            Q(source_region__name__icontains=query) |
            Q(source_region__code__icontains=query) |
            Q(source_area__name__icontains=query) |
            Q(source_area__code__icontains=query) |
            Q(source_branch__name__icontains=query) |
            Q(source_branch__code__icontains=query) |
            Q(source_zone__name__icontains=query) |
            Q(source_zone__code__icontains=query) |
            Q(source_department__name__icontains=query) |
            Q(source_department__code__icontains=query) |
            Q(destination_department__name__icontains=query) |
            Q(destination_department__code__icontains=query) |
            Q(current_department__name__icontains=query) |
            Q(current_department__code__icontains=query)
        )

    if status:
        documents = documents.filter(status=status)

    if priority:
        documents = documents.filter(priority=priority)

    if department_id:
        documents = documents.filter(
            Q(source_department_id=department_id) |
            Q(destination_department_id=department_id) |
            Q(current_department_id=department_id)
        )

    if zone_id:
        documents = documents.filter(source_zone_id=zone_id)

    if entry_type:
        documents = documents.filter(entry_type=entry_type)

    if date_from:
        documents = documents.filter(created_at__date__gte=date_from)

    if date_to:
        documents = documents.filter(created_at__date__lte=date_to)

    return documents

@login_required
def dashboard(request):
    pending_status_exclusions = ['closed', 'forwarded']
    my_assigned_documents_count = Document.objects.filter(
        assigned_documents_filter(request.user)
    ).exclude(status='closed').distinct().count()
    can_view_system_dashboard = can_view_system_documents(request.user)
    delegated_to_me_count = DocumentDelegation.objects.filter(
        delegated_recipient=request.user,
        is_active=True,
    ).exclude(document__status='closed').count()
    delegated_by_me_count = DocumentDelegation.objects.filter(
        Q(delegated_by=request.user) |
        Q(original_recipient=request.user),
        is_active=True,
    ).count()
    active_delegations_count = DocumentDelegation.objects.filter(is_active=True).count()
    active_leave_rules_count = UserDelegationRule.objects.filter(is_active=True).count()

    related_fields = [
        'source_zone', 'source_division', 'source_region', 'source_area',
        'source_branch', 'source_department', 'destination_department',
        'current_department', 'designated_person', 'first_boss', 'courier_rate',
    ]

    if can_view_system_dashboard:
        total_documents = Document.objects.count()
        pending_documents = Document.objects.exclude(status__in=pending_status_exclusions).count()
        forwarded_documents = Document.objects.filter(status='forwarded').count()
        closed_documents = Document.objects.filter(status='closed').count()
        urgent_pending = Document.objects.filter(
            priority='urgent'
        ).exclude(status__in=pending_status_exclusions).count()
        recent_documents = Document.objects.select_related(*related_fields).order_by('-created_at')[:8]
        dashboard_table_title = 'Recent Documents'
    else:
        assigned_documents = Document.objects.filter(
            assigned_documents_filter(request.user),
        ).distinct()
        total_documents = assigned_documents.count()
        closed_documents = assigned_documents.filter(status='closed').count()
        urgent_pending = assigned_documents.filter(
            priority='urgent'
        ).exclude(status='closed').count()
        recent_documents = assigned_documents.select_related(*related_fields).annotate(
            closed_sort=Case(
                When(status='closed', then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            )
        ).order_by('closed_sort', '-updated_at', '-created_at')[:8]
        dashboard_table_title = 'My Assigned Documents'

    context = {
        'total_documents': total_documents,
        'closed_documents': closed_documents,
        'urgent_pending': urgent_pending,
        'urgent_documents': urgent_pending,
        'recent_documents': recent_documents,
        'can_edit': can_edit_documents(request.user),
        'can_view_reports': can_view_reports(request.user),
        'can_view_system_dashboard': can_view_system_dashboard,
        'dashboard_table_title': dashboard_table_title,
        'my_assigned_documents_count': my_assigned_documents_count,
        'delegated_to_me_count': delegated_to_me_count,
        'delegated_by_me_count': delegated_by_me_count,
        'active_delegations_count': active_delegations_count,
        'active_leave_rules_count': active_leave_rules_count,
        'can_view_delegation_reports': can_view_delegation_reports(request.user),
    }
    if can_view_system_dashboard:
        context.update({
            'pending_documents': pending_documents,
            'forwarded_documents': forwarded_documents,
        })
    return render(request, 'documents/dashboard.html', context)


@login_required
def document_list(request):
    base_queryset = None
    if not can_view_system_documents(request.user):
        base_queryset = Document.objects.select_related(
            'source_zone',
            'source_division',
            'source_region',
            'source_area',
            'source_branch',
            'source_department',
            'destination_department',
            'current_department',
            'created_by',
            'designated_person',
            'first_boss',
            'courier_rate',
        ).filter(
            assigned_documents_filter(request.user),
        ).distinct().order_by('-created_at')

    documents = filter_documents_from_request(request, base_queryset)

    query = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()
    priority = request.GET.get('priority', '').strip()
    department_id = request.GET.get('department', '').strip()
    zone_id = request.GET.get('zone', '').strip()
    entry_type = request.GET.get('entry_type', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    departments = Department.objects.filter(is_active=True).order_by('name')
    zones = Zone.objects.filter(is_active=True).order_by('name')

    context = {
        'documents': documents,
        'departments': departments,
        'zones': zones,
        'query': query,
        'selected_status': status,
        'selected_priority': priority,
        'selected_department': department_id,
        'selected_zone': zone_id,
        'selected_entry_type': entry_type,
        'selected_date_from': date_from,
        'selected_date_to': date_to,
        'status_choices': Document.STATUS_CHOICES,
        'priority_choices': Document.PRIORITY_CHOICES,
        'entry_type_choices': Document.ENTRY_TYPE_CHOICES,
        'total_found': documents.count(),
        'can_edit': can_edit_documents(request.user),
    }
    return render(request, 'documents/document_list.html', context)

@login_required
def pending_documents(request):
    base_queryset = Document.objects.select_related(
        'source_zone',
        'source_division',
        'source_region',
        'source_area',
        'source_branch',
        'source_department',
        'destination_department',
        'current_department',
        'created_by',
        'designated_person',
        'first_boss',
        'courier_rate',
    ).exclude(status__in=['closed', 'forwarded']).order_by('created_at')

    if not can_view_system_documents(request.user):
        base_queryset = base_queryset.filter(
            assigned_documents_filter(request.user),
        ).distinct()

    documents = filter_documents_from_request(request, base_queryset)

    query = request.GET.get('q', '').strip()
    department_id = request.GET.get('department', '').strip()
    zone_id = request.GET.get('zone', '').strip()
    priority = request.GET.get('priority', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    departments = Department.objects.filter(is_active=True).order_by('name')
    zones = Zone.objects.filter(is_active=True).order_by('name')

    context = {
        'documents': documents,
        'departments': departments,
        'zones': zones,
        'query': query,
        'selected_department': department_id,
        'selected_zone': zone_id,
        'selected_priority': priority,
        'selected_date_from': date_from,
        'selected_date_to': date_to,
        'priority_choices': Document.PRIORITY_CHOICES,
        'total_found': documents.count(),
        'can_edit': can_edit_documents(request.user),
        'page_title': 'Pending Documents',
        'page_heading': 'Pending Documents',
        'page_description': 'Documents that are not forwarded or closed yet are listed here.',
        'records_title': 'Pending Records',
        'clear_url_name': 'pending_documents',
        'export_query_prefix': 'type=pending&',
    }
    return render(request, 'documents/pending_documents.html', context)


@login_required
def forwarded_documents(request):
    base_queryset = Document.objects.select_related(
        'source_zone',
        'source_division',
        'source_region',
        'source_area',
        'source_branch',
        'source_department',
        'destination_department',
        'current_department',
        'created_by',
        'designated_person',
        'first_boss',
        'courier_rate',
    ).filter(status='forwarded').order_by('-updated_at')

    if not can_view_system_documents(request.user):
        base_queryset = base_queryset.filter(
            assigned_documents_filter(request.user),
        ).distinct()

    documents = filter_documents_from_request(request, base_queryset)

    query = request.GET.get('q', '').strip()
    department_id = request.GET.get('department', '').strip()
    zone_id = request.GET.get('zone', '').strip()
    priority = request.GET.get('priority', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    departments = Department.objects.filter(is_active=True).order_by('name')
    zones = Zone.objects.filter(is_active=True).order_by('name')

    context = {
        'documents': documents,
        'departments': departments,
        'zones': zones,
        'query': query,
        'selected_department': department_id,
        'selected_zone': zone_id,
        'selected_priority': priority,
        'selected_date_from': date_from,
        'selected_date_to': date_to,
        'priority_choices': Document.PRIORITY_CHOICES,
        'total_found': documents.count(),
        'can_edit': can_edit_documents(request.user),
        'page_title': 'Forwarded Documents',
        'page_heading': 'Forwarded Documents',
        'page_description': 'Documents currently marked as forwarded are listed here.',
        'records_title': 'Forwarded Records',
        'clear_url_name': 'forwarded_documents',
        'export_query_prefix': 'status=forwarded&',
    }
    return render(request, 'documents/pending_documents.html', context)


@login_required
def my_assigned_documents(request):
    query = request.GET.get('q', '').strip()
    documents = Document.objects.select_related(
        'source_zone',
        'source_division',
        'source_region',
        'source_area',
        'source_branch',
        'source_department',
        'destination_department',
        'current_department',
        'created_by',
        'designated_person',
        'first_boss',
        'courier_rate',
    ).filter(
        assigned_documents_filter(request.user),
    ).exclude(status='closed').distinct().order_by('created_at')

    if query:
        documents = documents.filter(
            Q(tracking_id__icontains=query) |
            Q(reference_no__icontains=query) |
            Q(subject__icontains=query) |
            Q(document_type__icontains=query) |
            Q(remarks__icontains=query)
        )

    context = {
        'documents': documents,
        'query': query,
        'total_found': documents.count(),
    }
    return render(request, 'documents/my_assigned_documents.html', context)


@login_required
def bulk_forward_documents(request):
    if not can_bulk_forward(request.user):
        return permission_denied_page(
            request,
            'You do not have permission to bulk forward documents.'
        )

    if request.method != 'POST':
        messages.error(request, 'Bulk forwarding requires selected documents.')
        return redirect('document_list')

    next_url = safe_next_url(request)
    document_ids = selected_document_ids(request)
    if not document_ids:
        messages.warning(request, 'Please select at least one document to forward.')
        return redirect(next_url)

    documents = list(selected_documents_queryset(document_ids))
    actionable, skipped = partition_bulk_forward_documents(
        documents,
        requested_count=len(document_ids),
    )

    if not actionable:
        messages.warning(request, 'No selected documents can be forwarded.')
        return redirect(next_url)

    if request.POST.get('confirm') != '1':
        return render_bulk_confirmation(
            request,
            'forward',
            actionable,
            skipped,
        )

    remarks = request.POST.get('remarks', '').strip() or 'Bulk forwarded'
    forwarded_count = 0
    notification_failures = 0
    skipped_count = posted_int(request, 'skipped_count')

    for document in actionable:
        if document.status == 'closed':
            continue

        from_department = document.current_department
        to_department = document.destination_department or document.current_department

        with transaction.atomic():
            document.status = 'forwarded'
            if document.destination_department:
                document.current_department = document.destination_department
            document.save()

            DocumentMovement.objects.create(
                document=document,
                action='forwarded',
                from_department=from_department,
                to_department=to_department,
                performed_by=request.user,
                remarks=remarks,
            )

        notification_result = send_forward_notification(document, request.user)
        if notification_result in {'failed', 'missing_recipient'}:
            notification_failures += 1

        forwarded_count += 1

    messages.success(
        request,
        (
            f'{forwarded_count} documents forwarded successfully. '
            f'{notification_failures} notification failures.'
        )
    )
    if skipped_count:
        messages.warning(request, f'{skipped_count} selected documents were skipped.')
    return redirect(next_url)


@login_required
def bulk_close_documents(request):
    if not can_bulk_close(request.user):
        return permission_denied_page(
            request,
            'You do not have permission to bulk close documents.'
        )

    if request.method != 'POST':
        messages.error(request, 'Bulk closing requires selected documents.')
        return redirect('my_assigned_documents')

    next_url = safe_next_url(request, 'my_assigned_documents')
    document_ids = selected_document_ids(request)
    if not document_ids:
        messages.warning(request, 'Please select at least one document to close.')
        return redirect(next_url)

    documents = list(selected_documents_queryset(document_ids))
    actionable, skipped = partition_bulk_close_documents(
        request.user,
        documents,
        requested_count=len(document_ids),
    )

    if not actionable:
        messages.warning(request, 'No selected documents can be closed.')
        return redirect(next_url)

    if request.POST.get('confirm') != '1':
        return render_bulk_confirmation(
            request,
            'close',
            actionable,
            skipped,
        )

    remarks = request.POST.get('remarks', '').strip()
    if not remarks:
        return render_bulk_confirmation(
            request,
            'close',
            actionable,
            skipped,
            remarks_error='Closing remarks are required.',
        )

    closed_count = 0
    skipped_count = posted_int(request, 'skipped_count') + len(skipped)

    for document in actionable:
        if not can_user_close_document(request.user, document):
            skipped_count += 1
            continue

        from_department = document.current_department
        with transaction.atomic():
            document.status = 'closed'
            document.save()

            DocumentMovement.objects.create(
                document=document,
                action='closed',
                from_department=from_department,
                to_department=document.current_department,
                performed_by=request.user,
                remarks=remarks,
            )

        closed_count += 1

    messages.success(
        request,
        f'{closed_count} documents closed successfully. {skipped_count} skipped due to permission/status.'
    )
    return redirect(next_url)


@login_required
def document_create(request):
    if not can_edit_documents(request.user):
        return permission_denied_page(request, 'You do not have permission to create documents.')

    receiving_department = receiving_desk_department()

    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)

        if form.is_valid():
            document = form.save(commit=False)
            document.created_by = request.user
            document.save()

            DocumentMovement.objects.create(
                document=document,
                action='created',
                from_department=None,
                to_department=document.current_department,
                performed_by=request.user,
                remarks='Initial document entry created.'
            )

            return redirect('document_detail', pk=document.pk)
    else:
        initial = {}
        if receiving_department:
            initial['current_department'] = receiving_department.pk
        form = DocumentForm(initial=initial)

    context = {
        'form': form,
        **document_form_department_context(request),
    }
    return render(request, 'documents/document_form.html', context)

@login_required
def document_edit(request, pk):
    if not can_edit_documents(request.user):
        return permission_denied_page(request, 'You do not have permission to edit documents.')

    document = get_object_or_404(Document, pk=pk)

    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES, instance=document)

        if form.is_valid():
            old_department = document.current_department
            form.save()
            new_department = document.current_department

            if new_department and new_department != old_department:
                DocumentMovement.objects.create(
                    document=document,
                    action='forwarded',
                    from_department=old_department,
                    to_department=new_department,
                    performed_by=request.user,
                    remarks='Document forwarded to designated person\'s department.',
                )
            else:
                DocumentMovement.objects.create(
                    document=document,
                    action='status_updated',
                    from_department=old_department,
                    to_department=old_department,
                    performed_by=request.user,
                    remarks='Document information edited.'
                )

            return redirect('document_detail', pk=document.pk)
    else:
        form = DocumentForm(instance=document)

    context = {
        'form': form,
        'document': document,
        'is_edit': True,
        **document_form_department_context(request),
    }
    return render(request, 'documents/document_form.html', context)


@login_required
def document_detail(request, pk):
    document = get_object_or_404(
        Document.objects.select_related(
            'source_zone',
            'source_division',
            'source_region',
            'source_area',
            'source_branch',
        'source_department',
        'destination_department',
        'current_department',
        'created_by',
        'designated_person',
        'first_boss',
        'courier_rate',
        ).prefetch_related('delegations'),
        pk=pk
    )

    if not can_view_document(request.user, document):
        return permission_denied_page(request, 'You do not have permission to view this document.')

    active_delegation = active_delegation_for_document(document)
    movements = document.movements.select_related(
        'from_department',
        'to_department',
        'performed_by',
    ).all()

    context = {
        'document': document,
        'movements': movements,
        'can_edit': can_edit_documents(request.user),
        'can_update_status': can_update_document_status(request.user, document),
        'can_delegate': can_delegate_document(request.user, document),
        'active_delegation': active_delegation,
    }
    return render(request, 'documents/document_detail.html', context)


@login_required
def document_delegate_receipt(request, pk):
    document = get_object_or_404(
        Document.objects.select_related(
            'designated_person',
            'current_department',
            'destination_department',
        ).prefetch_related('delegations'),
        pk=pk,
    )

    if not is_designated_person_for_document(request.user, document):
        return permission_denied_page(request, 'Only the original designated recipient can delegate receipt.')

    if document.status == 'closed':
        return permission_denied_page(request, 'Closed documents cannot be delegated.')

    if active_delegation_for_document(document):
        return permission_denied_page(request, 'This document already has an active delegation.')

    temporary_rule = active_temporary_delegation_rule_for_user(request.user)
    initial = {}
    if temporary_rule:
        initial['delegated_recipient'] = temporary_rule.backup_receiver_id

    if request.method == 'POST':
        form = DocumentDelegationForm(request.POST, document=document)
        if form.is_valid():
            delegated_recipient = form.cleaned_data['delegated_recipient']
            reason = form.cleaned_data['reason']

            with transaction.atomic():
                if document.delegations.filter(is_active=True).exists():
                    form.add_error(None, 'This document already has an active delegation.')
                else:
                    delegation = DocumentDelegation.objects.create(
                        document=document,
                        original_recipient=document.designated_person,
                        delegated_recipient=delegated_recipient,
                        reason=reason,
                        delegated_by=request.user,
                        delegated_at=timezone.now(),
                    )
                    original_name = (
                        delegation.original_recipient.get_full_name() or
                        delegation.original_recipient.username
                    )
                    delegated_name = (
                        delegation.delegated_recipient.get_full_name() or
                        delegation.delegated_recipient.username
                    )
                    DocumentMovement.objects.create(
                        document=document,
                        action='delegated',
                        from_department=document.current_department,
                        to_department=document.current_department,
                        performed_by=request.user,
                        remarks=(
                            f'Receipt delegated from {original_name} to '
                            f'{delegated_name}. Reason: {reason}'
                        ),
                    )
                    messages.success(request, 'Receipt delegated successfully.')
                    return redirect('document_detail', pk=document.pk)
    else:
        form = DocumentDelegationForm(document=document, initial=initial)

    context = {
        'document': document,
        'form': form,
        'temporary_rule': temporary_rule,
    }
    return render(request, 'documents/document_delegate_form.html', context)


@login_required
def document_update_movement(request, pk):
    document = get_object_or_404(
        Document.objects.select_related(
            'current_department',
            'destination_department',
            'designated_person',
            'first_boss',
            'courier_rate',
        ),
        pk=pk
    )

    is_editor = can_edit_documents(request.user)
    can_update_status = can_update_document_status(request.user, document)

    if not can_update_status:
        return permission_denied_page(request, 'You do not have permission to update or forward documents.')

    status_choices = None if is_editor else [('closed', 'Closed')]
    allow_department_change = is_editor

    if request.method == 'POST':
        form = DocumentMovementForm(
            request.POST,
            status_choices=status_choices,
            allow_department_change=allow_department_change,
        )

        if form.is_valid():
            new_status = form.cleaned_data['new_status']
            to_department = form.cleaned_data.get('to_department')

            if not is_editor:
                to_department = None

            if not is_editor and new_status != 'closed':
                form.add_error(
                    'new_status',
                    'Assigned users can only close their assigned document.'
                )

            elif (
                new_status == 'forwarded' and
                document.notification_required and
                not document.notification_sent and
                (
                    not document.designated_person or
                    not document.designated_person.email
                )
            ):
                form.add_error(
                    None,
                    'Forwarding requires a designated person with an email when notification is required.'
                )
            else:
                movement = form.save(commit=False)
                movement.document = document
                movement.action = movement_action_from_status(new_status)
                movement.from_department = document.current_department
                movement.performed_by = request.user
                movement.save()

                document.status = new_status

                if to_department:
                    document.current_department = to_department

                if new_status == 'forwarded' and to_department:
                    document.destination_department = to_department

                document.save()

                if new_status == 'forwarded':
                    send_forward_notification(document, request.user)

                return redirect('document_detail', pk=document.pk)
    else:
        form = DocumentMovementForm(
            initial={
                'new_status': 'closed' if not is_editor else document.status,
                'to_department': document.current_department,
            },
            status_choices=status_choices,
            allow_department_change=allow_department_change,
        )

    context = {
        'document': document,
        'form': form,
        'is_editor_update': is_editor,
    }
    return render(request, 'documents/document_movement_form.html', context)


@login_required
def my_delegation_settings(request):
    if request.method == 'POST':
        form = UserDelegationRuleForm(request.POST, user=request.user)
        if form.is_valid():
            rule = form.save(commit=False)
            rule.user = request.user
            rule.created_by = request.user
            rule.save()
            messages.success(request, 'Temporary delegation rule saved.')
            return redirect('my_delegation_settings')
    else:
        form = UserDelegationRuleForm(user=request.user)

    today = timezone.now().date()
    rules = UserDelegationRule.objects.filter(user=request.user).select_related(
        'backup_receiver',
        'created_by',
    ).order_by('-is_active', '-start_date', '-created_at')
    active_rule = rules.filter(
        is_active=True,
        start_date__lte=today,
        end_date__gte=today,
    ).first()

    context = {
        'form': form,
        'rules': rules,
        'active_rule': active_rule,
    }
    return render(request, 'documents/my_delegation_settings.html', context)


@login_required
def reports_dashboard(request):
    if not can_view_reports(request.user):
        return permission_denied_page(request, 'You do not have permission to view reports.')

    today = timezone.now().date()
    closed_status = 'closed'
    forwarded_status = 'forwarded'
    urgent_priority = 'urgent'
    pending_status_exclusions = [closed_status, forwarded_status]

    total_documents = Document.objects.count()
    pending_documents = Document.objects.exclude(status__in=pending_status_exclusions).count()
    forwarded_documents = Document.objects.filter(status=forwarded_status).count()
    closed_documents = Document.objects.filter(status=closed_status).count()
    urgent_open = Document.objects.filter(priority=urgent_priority).exclude(status=closed_status).count()

    received_today = Document.objects.filter(received_date=today).count()
    forwarded_today = DocumentMovement.objects.filter(
        action=forwarded_status,
        created_at__date=today,
    ).values('document_id').distinct().count()
    if not DocumentMovement.objects.filter(action=forwarded_status).exists():
        forwarded_today = Document.objects.filter(
            status=forwarded_status,
            updated_at__date=today,
        ).count()
    closed_today = Document.objects.filter(
        Q(closed_at__date=today) |
        Q(status=closed_status, closed_at__isnull=True, updated_at__date=today)
    ).count()
    notifications_sent = Document.objects.filter(notification_sent=True).count()
    notifications_failed = Document.objects.exclude(notification_error='').count()
    active_delegations = DocumentDelegation.objects.filter(is_active=True).count()
    active_leave_rules = UserDelegationRule.objects.filter(is_active=True).count()
    total_outwarding_documents = Document.objects.filter(entry_type='outward').count()
    total_dispatch_cost = (
        Document.objects.filter(entry_type='outward').aggregate(
            total=Sum('outward_amount')
        )['total'] or 0
    )

    department_pending = Department.objects.filter(is_active=True).annotate(
        pending_count=Count(
            'current_documents',
            filter=~Q(current_documents__status__in=pending_status_exclusions)
        )
    ).order_by('-pending_count', 'name')

    department_forwarded = Department.objects.filter(is_active=True).annotate(
        forwarded_count=Count(
            'current_documents',
            filter=Q(current_documents__status=forwarded_status)
        )
    ).order_by('-forwarded_count', 'name')

    raw_division_documents = Document.objects.values(
        'source_division__name'
    ).annotate(
        document_count=Count('id')
    ).order_by('-document_count', 'source_division__name')

    division_documents = [
        {
            'name': item['source_division__name'] or 'External / Not Specified',
            'document_count': item['document_count'],
        }
        for item in raw_division_documents
    ]

    aging_documents = Document.objects.select_related(
        'source_zone',
        'source_division',
        'source_region',
        'source_area',
        'source_branch',
        'source_department',
        'destination_department',
        'current_department',
        'designated_person',
        'first_boss',
        'courier_rate',
    ).exclude(status=closed_status).order_by('created_at')[:10]

    recent_closed_documents = Document.objects.select_related(
        'source_zone',
        'source_division',
        'source_region',
        'source_area',
        'source_branch',
        'source_department',
        'destination_department',
        'current_department',
        'designated_person',
        'first_boss',
        'courier_rate',
    ).filter(status=closed_status).order_by('-closed_at')[:10]

    failed_notification_documents = Document.objects.select_related(
        'designated_person',
        'first_boss',
        'current_department',
        'courier_rate',
    ).exclude(notification_error='').order_by('-updated_at')[:10]

    recent_outwarding_cost_documents = Document.objects.select_related(
        'courier_rate',
        'current_department',
    ).filter(
        entry_type='outward',
        outward_amount__isnull=False,
    ).order_by('-created_at')[:10]

    raw_status_counts = Document.objects.values('status').annotate(
        count=Count('id')
    )
    status_counts = {
        item['status']: item['count']
        for item in raw_status_counts
    }

    status_summary = [
        {
            'label': label,
            'count': status_counts.get(value, 0),
        }
        for value, label in Document.STATUS_CHOICES
    ]

    raw_entry_type_counts = Document.objects.values('entry_type').annotate(
        count=Count('id')
    )
    entry_type_counts = {
        item['entry_type']: item['count']
        for item in raw_entry_type_counts
    }

    entry_type_summary = [
        {
            'label': label,
            'count': entry_type_counts.get(value, 0),
        }
        for value, label in Document.ENTRY_TYPE_CHOICES
    ]

    context = {
        'total_documents': total_documents,
        'pending_documents': pending_documents,
        'forwarded_documents': forwarded_documents,
        'closed_documents': closed_documents,
        'urgent_open': urgent_open,
        'received_today': received_today,
        'forwarded_today': forwarded_today,
        'closed_today': closed_today,
        'notifications_sent': notifications_sent,
        'notifications_failed': notifications_failed,
        'active_delegations': active_delegations,
        'active_leave_rules': active_leave_rules,
        'can_view_delegation_reports': can_view_delegation_reports(request.user),
        'total_outwarding_documents': total_outwarding_documents,
        'total_dispatch_cost': total_dispatch_cost,
        'department_pending': department_pending,
        'department_forwarded': department_forwarded,
        'division_documents': division_documents,
        'aging_documents': aging_documents,
        'recent_closed_documents': recent_closed_documents,
        'failed_notification_documents': failed_notification_documents,
        'recent_outwarding_cost_documents': recent_outwarding_cost_documents,
        'status_summary': status_summary,
        'entry_type_summary': entry_type_summary,
    }

    return render(request, 'documents/reports_dashboard.html', context)


@login_required
def delegation_report(request):
    if not can_view_delegation_reports(request.user):
        return permission_denied_page(request, 'You do not have permission to view delegation reports.')

    user_model = get_user_model()
    original_recipient_id = request.GET.get('original_recipient', '').strip()
    delegated_recipient_id = request.GET.get('delegated_recipient', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    status = request.GET.get('status', '').strip()
    document_status = request.GET.get('document_status', '').strip()

    delegations = DocumentDelegation.objects.select_related(
        'document',
        'original_recipient',
        'delegated_recipient',
        'delegated_by',
        'cancelled_by',
    ).order_by('-delegated_at')

    if original_recipient_id.isdigit():
        delegations = delegations.filter(original_recipient_id=original_recipient_id)

    if delegated_recipient_id.isdigit():
        delegations = delegations.filter(delegated_recipient_id=delegated_recipient_id)

    if date_from:
        delegations = delegations.filter(delegated_at__date__gte=date_from)

    if date_to:
        delegations = delegations.filter(delegated_at__date__lte=date_to)

    if status == 'active':
        delegations = delegations.filter(is_active=True)
    elif status == 'cancelled':
        delegations = delegations.filter(is_active=False)

    if document_status:
        delegations = delegations.filter(document__status=document_status)

    users = user_model.objects.filter(is_active=True).order_by(
        'first_name',
        'last_name',
        'username',
    )
    report_rows = []
    for delegation in delegations:
        closed_movement = delegation.document.movements.filter(
            action='closed',
        ).select_related('performed_by').order_by('-created_at').first()
        report_rows.append({
            'delegation': delegation,
            'closed_by': closed_movement.performed_by if closed_movement else None,
            'closed_at': delegation.document.closed_at or (
                closed_movement.created_at if closed_movement else None
            ),
        })

    context = {
        'report_rows': report_rows,
        'users': users,
        'status_choices': [
            ('active', 'Active'),
            ('cancelled', 'Cancelled'),
        ],
        'document_status_choices': Document.STATUS_CHOICES,
        'selected_original_recipient': original_recipient_id,
        'selected_delegated_recipient': delegated_recipient_id,
        'selected_date_from': date_from,
        'selected_date_to': date_to,
        'selected_status': status,
        'selected_document_status': document_status,
        'total_found': len(report_rows),
    }
    return render(request, 'documents/delegation_report.html', context)


@login_required
def export_documents_excel(request):
    if not can_view_reports(request.user):
        return permission_denied_page(request, 'You do not have permission to export documents.')

    export_type = request.GET.get('type', '').strip()

    if export_type == 'pending':
        base_queryset = Document.objects.select_related(
            'source_zone',
            'source_division',
            'source_region',
            'source_area',
            'source_branch',
            'source_department',
            'destination_department',
            'current_department',
            'created_by',
            'designated_person',
            'first_boss',
            'courier_rate',
        ).exclude(status__in=['closed', 'forwarded']).order_by('created_at')

        documents = filter_documents_from_request(request, base_queryset)
    else:
        documents = filter_documents_from_request(request)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Documents'

    headers = [
        'Tracking ID',
        'Reference No',
        'Entry Type',
        'Document Type',
        'Subject',
        'Addressed To Name',
        'Addressed To Designation',
        'Designated Person',
        'Designated Person Email',
        'First Boss',
        'First Boss Email',
        'Source Type',
        'External Organization Type',
        'External Organization Name',
        'External Branch',
        'Source Division',
        'Source Region',
        'Source Area',
        'Source Branch',
        'Source Department',
        'Destination Department',
        'Receiving Department',
        'Status',
        'Priority',
        'Notification Required',
        'Notification Sent',
        'Notification Sent At',
        'Notification Error',
        'Courier Description',
        'Courier Quantity',
        'Courier Amount',
        'Courier Rate Source',
        'Received Date',
        'Sent Date',
        'Aging Days',
        'Remarks',
        'Created By',
        'Created At',
    ]

    sheet.append(headers)

    header_fill = PatternFill(start_color='0F6B45', end_color='0F6B45', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for document in documents:
        sheet.append([
            document.tracking_id,
            document.reference_no,
            document.get_entry_type_display(),
            document.document_type,
            document.subject,
            document.addressed_to_name,
            document.addressed_to_designation,
            document.designated_person.username if document.designated_person else '',
            document.designated_person.email if document.designated_person else '',
            document.first_boss.username if document.first_boss else '',
            document.first_boss.email if document.first_boss else '',
            document.get_source_type_display(),
            document.get_external_organization_type_display() if document.external_organization_type else '',
            document.external_organization_name,
            document.external_branch_name,
            document.source_division.name if document.source_division else '',
            document.source_region.name if document.source_region else '',
            document.source_area.name if document.source_area else '',
            document.source_branch.name if document.source_branch else '',
            document.source_department.name if document.source_department else '',
            document.destination_department.name if document.destination_department else '',
            document.current_department.name if document.current_department else '',
            document.get_status_display(),
            document.get_priority_display(),
            'Yes' if document.notification_required else 'No',
            'Yes' if document.notification_sent else 'No',
            document.notification_sent_at.strftime('%Y-%m-%d %I:%M %p') if document.notification_sent_at else '',
            document.notification_error,
            document.outward_description,
            document.outward_quantity,
            document.outward_amount if document.outward_amount is not None else '',
            courier_rate_source(document),
            document.received_date.strftime('%Y-%m-%d') if document.received_date else '',
            document.sent_date.strftime('%Y-%m-%d') if document.sent_date else '',
            document.aging_days,
            document.remarks,
            document.created_by.username if document.created_by else '',
            document.created_at.strftime('%Y-%m-%d %I:%M %p'),
        ])

    column_widths = {
        'A': 18,
        'B': 18,
        'C': 14,
        'D': 18,
        'E': 35,
        'F': 24,
        'G': 24,
        'H': 22,
        'I': 28,
        'J': 22,
        'K': 28,
        'L': 14,
        'M': 26,
        'N': 28,
        'O': 24,
        'P': 20,
        'Q': 20,
        'R': 20,
        'S': 24,
        'T': 20,
        'U': 24,
        'V': 24,
        'W': 25,
        'X': 12,
        'Y': 20,
        'Z': 18,
        'AA': 20,
        'AB': 35,
        'AC': 28,
        'AD': 24,
        'AE': 18,
        'AF': 22,
        'AG': 15,
        'AH': 15,
        'AI': 12,
        'AJ': 35,
        'AK': 15,
        'AL': 20,
    }

    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="document_records.xlsx"'

    workbook.save(response)
    return response
